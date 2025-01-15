import datetime
import re
import time
from pathlib import Path

import dateutil
import pandas as pd
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.text import RichText
from openpyxl.drawing.fill import ColorChoice, PatternFillProperties
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.text import (CharacterProperties, Paragraph,
                                   ParagraphProperties)
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import (Alignment, Border, Font,
                             PatternFill, Side)
from openpyxl.styles.numbers import FORMAT_NUMBER_00 , FORMAT_PERCENTAGE_00
from openpyxl.utils.dataframe import dataframe_to_rows  # Per l'import di dataframe
from openpyxl.utils.units import cm_to_EMU, pixels_to_EMU
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.page import PageMargins  # Opzioni di stampa
from openpyxl.worksheet.worksheet import Worksheet


class Report():
    """Crea un report di un portafoglio."""

    def __init__(self, t1, file_portafoglio='artes.xlsx'):
        """
        Initialize the class.

        Arguments:
            t1 {str} = data finale
            file_portafoglio {str} = nome del file excel da lavorare
        """
        self.wb = Workbook()

        # Dates
        self.t1 = datetime.datetime.strptime(t1, '%d/%m/%Y')
        print(f"Data report : {self.t1}.")
        # t1 un mese fa
        self.t0_1m = self.t1.replace(day=1) - dateutil.relativedelta.relativedelta(days=1)
        print(f"Un mese fa : {self.t0_1m}.")
        # t1 un anno fa
        self.t0_1Y = (self.t1 - dateutil.relativedelta.relativedelta(years=+1))#.strftime("%d/%m/%Y") 
        print(f"Un anno fa : {self.t0_1Y}.")
        # t1 all fine dell'anno scorso
        self.t0_ytd = datetime.datetime(year=self.t0_1Y.year, month=12, day=31)
        print(f"L'ultimo giorno dell'anno scorso : {self.t0_ytd}.")
        # t1 tre anni fa
        self.t0_3Y = (self.t1 - dateutil.relativedelta.relativedelta(years=+3))#.strftime("%d/%m/%Y")
        # self.t0_3Y = '2020-02-29 00:00:00'
        print(f"Tre anni fa : {self.t0_3Y}.")

        # Directory
        directory = Path().cwd()
        self.path = directory
        self.file_portafoglio = self.path.joinpath(file_portafoglio)
        self.mesi_dict = {
            1: 'Gennaio', 2: 'Febbraio', 3: 'Marzo', 4: 'Aprile', 5: 'Maggio', 6: 'Giugno', 7: 'Luglio', 8: 'Agosto', 
            9: 'Settembre', 10: 'Ottobre', 11: 'Novembre', 12: 'Dicembre'
        }
        # Pipeline with Python and Postrge
        # DATABASE_URL = 'postgres+psycopg2://postgres:bloomberg893@localhost:5432/artes'
        # self.engine = create_engine(DATABASE_URL)
        # self.connection = self.engine.connect()

        # Carica portafoglio
        portfolio = pd.read_excel(self.file_portafoglio, sheet_name='Portfolio', header=1)

        # Controvalori
        controvalore_t1 = portfolio['TOTALE t1'].sum()
        print(f"\nIl controvalore del portafoglio è : {round(controvalore_t1, 2)}.")
        controvalore_t0_1m = portfolio['TOTALE t0'].sum()
        print(f"Il controvalore del portafoglio nel mese precedente era : {round(controvalore_t0_1m, 2)}.")

    def __logo(self, ws: Worksheet, picture: Path | str = Path(r'.\\img\\logo_B&S.bmp'),
        col: int = 5, colOff: float = 0.3, row: int = 34, rowOff: float = 0):   
        """
        Aggiunge un'immagine in coordinate precise del foglio, applicando uno spostamento.

        Arguments:
            ws {Worksheet} -- foglio in cui incollare l'immagine

        Keyword Arguments:
            picture {Path  |  str} -- percorso in cui si trova l'immagine (default: {Path(r'.\\img\\logo_B&S.bmp')})
            col {int} -- colonna di partenza in cui incollare l'immagine (default: {5})
            colOff {float} -- spostamento dalla colonna di partenza (default: {0.3})
            row {int} -- riga di partenza in cui incollare l'immagine (default: {34})
            rowOff {float} -- spostamento dalla riga di partenza (default: {0})      
        """
        logo = Image(picture)
        h, w = logo.height, logo.width
        size = XDRPositiveSize2D(pixels_to_EMU(w), pixels_to_EMU(h))
        cellw = lambda x: cm_to_EMU((x * (18.65-1.71))/10)
        cellh = lambda x: cm_to_EMU((x * 49.77)/99)
        maker = AnchorMarker(col=col, colOff=cellw(colOff), row=row, rowOff=cellh(rowOff))
        ancoraggio = OneCellAnchor(_from=maker, ext=size)
        ws.add_image(logo)
        logo.anchor = ancoraggio

    def __textbox(self, ws: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int):
        """Simulazione di una text-box

        Arguments:
            ws {Worksheet} -- foglio excel in cui creare la text box
            min_row {int} -- coordinate dove inserire la text box
            max_row {int} -- coordinate dove inserire la text box
            min_col {int} -- coordinate dove inserire la text box
            max_col {int} -- coordinate dove inserire la text box
        """        
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for _ in range(max_col-min_col+1):
                ws[row[_].coordinate].fill = PatternFill(fill_type='solid', fgColor='FFFFFF')
                ws[row[_].coordinate].font = Font(name='Times New Roman', size=12, color='31869B')
            ws[row[0].coordinate].border = Border(left=Side(border_style='medium', color='31869B'))
            ws[row[max_col-min_col].coordinate].border = Border(right=Side(border_style='medium', color='31869B'))
            if row[0].row == min_row:
                for _ in range(max_col-min_col+1):
                    if row[_].column == min_col:
                        ws[row[_].coordinate].border = Border(
                            top=Side(border_style='medium', color='31869B'), 
                            left=Side(border_style='medium', color='31869B'))
                    elif row[_].column == max_col:
                        ws[row[_].coordinate].border = Border(
                            top=Side(border_style='medium', color='31869B'), 
                            right=Side(border_style='medium', color='31869B'))
                    else:
                        ws[row[_].coordinate].border = Border(top=Side(border_style='medium', color='31869B'))
            elif row[0].row == max_row:
                for _ in range(max_col-min_col+1):
                    if row[_].column == min_col:
                        ws[row[_].coordinate].border = Border(
                            bottom=Side(border_style='medium', color='31869B'), 
                            left=Side(border_style='medium', color='31869B'))
                    elif row[_].column == max_col:
                        ws[row[_].coordinate].border = Border(
                            bottom=Side(border_style='medium', color='31869B'), 
                            right=Side(border_style='medium', color='31869B'))
                    else:
                        ws[row[_].coordinate].border = Border(
                            bottom=Side(border_style='medium', color='31869B'))

    def copertina_1(self):
        """
        Crea la copertina
        """
        ws = self.wb.active
        ws.title = '1.copertina'
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells('A1:L1')
        ws['A11'] = 'Benchmark & Style'
        ws['A11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A11'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A11'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells('A11:L14')
        ws['E17'] = 'ARTES'
        ws['E17'].alignment = Alignment(horizontal='center', vertical='center')
        ws['E17'].font = Font(name='Times New Roman', size=18, bold=True, color='31869B')
        ws['G17'] = self.t1.strftime('%d/%m/%Y')
        ws['G17'].alignment = Alignment(horizontal='center', vertical='center')
        ws['G17'].font = Font(name='Times New Roman', size=18, bold=True, color='31869B')
        ws.merge_cells('E17:F19')
        ws.merge_cells('G17:H19')
        ws['A33'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells('A33:L33')
        # Logo
        logo = Image(Path(r'.\\img\\logo_B&S.bmp'))
        ws.add_image(logo, 'F27')
        logo.height = 75.59
        logo.width = 128.88188976377952755905511811024

    def indice_2(self):
        """
        Crea la seconda pagina.
        """
        ws = self.wb.create_sheet('2.indice')
        ws = self.wb['2.indice']
        self.wb.active = ws
        ws.merge_cells('A1:L4')
        ws['A1'] = 'Indice'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws['B8'] = '1. Analisi Di Mercato'
        ws['B8'].font = Font(name='Times New Roman', size=18, bold=True, color='31869B')
        ws['B11'] = '2. Performance'
        ws['B11'].font = Font(name='Times New Roman', size=18, bold=True, color='31869B')
        ws['B14'] = '3. Valutazione Per Macroclasse'
        ws['B14'].font = Font(name='Times New Roman', size=18, bold=True, color='31869B')
        ws['B17'] = '4. Contatti'
        ws['B17'].font = Font(name='Times New Roman', size=18, bold=True, color='31869B')
        # Logo
        self.__logo(ws, row=32)
    
    def analisi_di_mercato_3(self):
        """
        Crea la terza pagina.
        """
        # 3.Analisi mercato
        ws = self.wb.create_sheet('3.an_mkt')
        ws = self.wb['3.an_mkt']
        self.wb.active = ws
        ws['A11'] = '1. Analisi Di Mercato'
        ws['A11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A11'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A11'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells('A11:L14')
        # Logo
        self.__logo(ws)

    def analisi_rendimenti_4(self):
        """
        Crea la quarta pagina.
        Aggiunge fogli Indici e fogli Indici_in_euro.
        """
        # Carica indici e tassi di cambio
        indici_tassi = pd.read_excel(self.file_portafoglio, sheet_name='Indici', header=[0, 1], parse_dates=True, index_col=0)
        indici_tassi.columns = indici_tassi.columns.droplevel(-1)
        # print(indici_tassi)
        # Carica indici in euro
        indici_in_euro = pd.read_excel(self.file_portafoglio, sheet_name='Indici_in_euro', header=[0, 1], parse_dates=True, index_col=0)
        indici_in_euro.columns = indici_in_euro.columns.droplevel(-1)
        # print(indici_in_euro)
        # Crea dizionario dove inserire i rendimenti degli indici e dei tassi
        indici_perf = {
            'S&P 500' : [], 'NIKKEI' : [], 'NASDAQ' : [], 'FTSE 100' : [], 'FTSE MIB' : [], 'DAX' : [],
            'DOW JONES INDUSTRIAL AVERAGE' : [], 'EURO STOXX 50' : [], 'HANG SENG' : [], 'MSCI WORLD' : [],
            'MSCI EMERGING MARKETS' : [], 'HFRX EWSI' : [], 'WTI CRUDE OIL FUTURE' : [], 'LONDON GOLD MARKET FIXING LTD' : [],
            'COMMODITY RESEARCH BUREAU' : [], 'LYXOR ETF EURO CASH' : [], 'LYXOR ETF EURO CORP BOND' : [],
            'BARCLAYS EUROAGG CORP TR' : [], 'JPM GBI EMU 1_10' : [], 'JPM GBI EMU 3_5' : [], 'JPM GBI EMU 1_3' : [], 
            'USDEUR' : [], 'GBPEUR' : [], 'CHFEUR' : [], 'AUDEUR' : [], 'NOKEUR' : []
        }
        for column in indici_tassi.columns.values:
            monthly = (indici_tassi.loc[self.t1, column] - indici_tassi.loc[self.t0_1m, column]) / (indici_tassi.loc[self.t0_1m, column])
            ytd = (indici_tassi.loc[self.t1, column] - indici_tassi.loc[self.t0_ytd, column]) / (indici_tassi.loc[self.t0_ytd, column])
            one_year = (indici_tassi.loc[self.t1, column] - indici_tassi.loc[self.t0_1Y, column]) / (indici_tassi.loc[self.t0_1Y, column])
            three_years = (indici_tassi.loc[self.t1, column] - indici_tassi.loc[self.t0_3Y, column]) / (indici_tassi.loc[self.t0_3Y, column])
            indici_perf[column] = [monthly, ytd, one_year, three_years]
        for column in indici_in_euro.columns.values:
            ytd_eur = (indici_in_euro.loc[self.t1, column] - indici_in_euro.loc[self.t0_ytd, column]) / (indici_in_euro.loc[self.t0_ytd, column])
            indici_perf[column].append(ytd_eur)
        # print(indici_perf)

        ws = self.wb.create_sheet('4.an_mkt_rend')
        ws = self.wb['4.an_mkt_rend']
        self.wb.active = ws
        ws.merge_cells('A1:O4')
        ws['A1'] = 'Analisi Di Mercato'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.row_dimensions[5].height = 3
        ws['A6'] = 'Performance ' + self.mesi_dict[self.t1.month]
        ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A6'].font = Font(name='Times New Roman', size=10, bold=True, color='FFFFFF')
        ws['A6'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws['A6'].border = Border(top=Side(border_style='medium', color='000000'), bottom=Side(border_style='medium', color='000000'), right=Side(border_style='medium', color='000000'), left=Side(border_style='medium', color='000000'))
        ws['A7'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells('A6:I6')

        # Colonne tabella
        header_4 = ['', '', '', '', 'MENSILI', 'YTD €', 'YTD', '1y', '3y']
        for column in ws.iter_cols(min_row=7, max_row=7, min_col=1, max_col=9):
            ws[column[0].coordinate].value = header_4[0]
            del header_4[0]
            ws[column[0].coordinate].font = Font(name='Times New Roman', size=9, bold=True, color='FFFFFF')
            ws[column[0].coordinate].fill = PatternFill(fill_type='solid', fgColor='31869B')
            ws[column[0].coordinate].alignment = Alignment(horizontal='center', vertical='center')
            ws[column[0].coordinate].border = Border(top=Side(border_style='medium', color='000000'), bottom=Side(border_style='medium', color='000000'))
            ws.cell(row=column[0].row, column=1).border = Border(left=Side(border_style='medium', color='000000'))
            ws.cell(row=column[0].row, column=9).border = Border(right=Side(border_style='medium', color='000000'))

        # Corpo tabella
        index_4 = [
            'AZIONARI', 'S&P 500', 'NIKKEI', 'NASDAQ', 'FTSE 100', 'FTSE MIB', 'DAX', 'DOW JONES INDUSTRIAL AVERAGE', 'EURO STOXX 50',
            'HANG SENG', 'MSCI WORLD', 'MSCI EMERGING MARKETS', 'HEDGE FUND', 'HFRX EWSI', 'COMMODITIES', 'WTI CRUDE OIL FUTURE',
            'LONDON GOLD MARKET FIXING LTD', 'COMMODITY RESEARCH BUREAU', 'OBBLIGAZIONARI GOVERNATIVE', 'LYXOR ETF EURO CASH',
            'JPM GBI EMU 1_3', 'JPM GBI EMU 3_5', 'JPM GBI EMU 1_10', 'OBBLIGAZIONARI CORPORATE', 'LYXOR ETF EURO CORP BOND',
            'BARCLAYS EUROAGG CORP TR', 'VALUTE', 'USDEUR', 'GBPEUR', 'CHFEUR', 'AUDEUR', 'NOKEUR',
            "le valute sono espresse come quantità di euro per un'unità di valuta estera"
        ]
        for row in ws.iter_rows(min_row=8, max_row=40, min_col=1, max_col=9):
            ws[row[0].coordinate].value = index_4[0]
            del index_4[0]
            ws.row_dimensions[row[0].row].height = 13
            if ws[row[0].coordinate].value == 'AZIONARI' or ws[row[0].coordinate].value == 'HEDGE FUND' or ws[row[0].coordinate].value == 'COMMODITIES'	or ws[row[0].coordinate].value == 'OBBLIGAZIONARI GOVERNATIVE' or ws[row[0].coordinate].value == 'OBBLIGAZIONARI CORPORATE' or ws[row[0].coordinate].value == 'VALUTE' or ws[row[0].coordinate].value == 'le valute sono espresse come quantità di euro per un\'unità di valuta estera':
                ws[row[0].coordinate].font = Font(name='Times New Roman', size=8, bold=True, color='006666')
                ws[row[0].coordinate].fill = PatternFill(fill_type='solid', fgColor='92CDDC')
                ws[row[0].coordinate].alignment = Alignment(horizontal='center', vertical='center')
                ws[row[0].coordinate].border = Border(top=Side(border_style='medium', color='000000'), left=Side(border_style='medium', color='000000'), bottom=Side(border_style='medium', color='000000'), right=Side(border_style='medium', color='000000'))
                ws.merge_cells(start_row=row[0].row, end_row=row[0].row, start_column=1, end_column=9)
            else:
                ws[row[0].coordinate].font = Font(name='Times New Roman', size=8, bold=True, color='FFFFFF')
                ws[row[0].coordinate].fill = PatternFill(fill_type='solid', fgColor='31869B')
                ws[row[0].coordinate].alignment = Alignment(vertical='center')
                ws[row[0].coordinate].border = Border(left=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'))
                ws.merge_cells(start_row=row[0].row, end_row=row[0].row, start_column=1, end_column=4)
            if ws[row[0].coordinate].value in indici_perf.keys():
                if row[0].row < 34: # Riempi la tabella tranne valute (00B050 + FF0000 -)
                    ws[row[4].coordinate].value = "{0:.2f}%".format(round(indici_perf[row[0].value][0] * 100, 2)).replace('.', ',')
                    if indici_perf[row[0].value][0] > 0:
                        ws[row[4].coordinate].font = Font(color='00B050')
                    else:
                        ws[row[4].coordinate].font = Font(color='FF0000')
                    ws[row[4].coordinate].alignment = Alignment(horizontal='center', vertical='center')
                    ws[row[4].coordinate].border = Border(bottom=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'))
                    ws[row[5].coordinate].value = "{0:.2f}%".format(round(indici_perf[row[0].value][4] * 100, 2)).replace('.', ',')
                    if indici_perf[row[0].value][4] > 0:
                        ws[row[5].coordinate].font = Font(color='00B050')
                    else:
                        ws[row[5].coordinate].font = Font(color='FF0000')
                    ws[row[5].coordinate].alignment = Alignment(horizontal='center', vertical='center')
                    ws[row[5].coordinate].border = Border(bottom=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'))
                    ws[row[6].coordinate].value = "{0:.2f}%".format(round(indici_perf[row[0].value][1] * 100, 2)).replace('.', ',')
                    if indici_perf[row[0].value][1] > 0:
                        ws[row[6].coordinate].font = Font(color='00B050')
                    else:
                        ws[row[6].coordinate].font = Font(color='FF0000')
                    ws[row[6].coordinate].alignment = Alignment(horizontal='center', vertical='center')
                    ws[row[6].coordinate].border = Border(bottom=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'))
                    ws[row[7].coordinate].value = "{0:.2f}%".format(round(indici_perf[row[0].value][2] * 100, 2)).replace('.', ',')
                    if indici_perf[row[0].value][2] > 0:
                        ws[row[7].coordinate].font = Font(color='00B050')
                    else:
                        ws[row[7].coordinate].font = Font(color='FF0000')
                    ws[row[7].coordinate].alignment = Alignment(horizontal='center', vertical='center')
                    ws[row[7].coordinate].border = Border(bottom=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'))
                    ws[row[8].coordinate].value = "{0:.2f}%".format(round(indici_perf[row[0].value][3] * 100, 2)).replace('.', ',')
                    if indici_perf[row[0].value][3] > 0:
                        ws[row[8].coordinate].font = Font(color='00B050')
                    else:
                        ws[row[8].coordinate].font = Font(color='FF0000')
                    ws[row[8].coordinate].alignment = Alignment(horizontal='center', vertical='center')
                    ws[row[8].coordinate].border = Border(bottom=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'))
                else: # Riempi valute
                    ws[row[4].coordinate].value = "{0:.2f}%".format(round(indici_perf[row[0].value][0] * 100, 2)).replace('.', ',')
                    if indici_perf[row[0].value][0] > 0:
                        ws[row[4].coordinate].font = Font(color='00B050')
                    else:
                        ws[row[4].coordinate].font = Font(color='FF0000')
                    ws[row[4].coordinate].alignment = Alignment(horizontal='center', vertical='center')
                    ws[row[4].coordinate].border = Border(bottom=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'))
                    ws[row[5].coordinate].value = "{0:.2f}%".format(round(indici_perf[row[0].value][1] * 100, 2)).replace('.', ',')
                    if indici_perf[row[0].value][1] > 0:
                        ws[row[5].coordinate].font = Font(color='00B050')
                    else:
                        ws[row[5].coordinate].font = Font(color='FF0000')
                    ws[row[5].coordinate].alignment = Alignment(horizontal='center', vertical='center')
                    ws[row[5].coordinate].border = Border(bottom=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'))
                    ws[row[7].coordinate].value = "{0:.2f}%".format(round(indici_perf[row[0].value][2] * 100, 2)).replace('.', ',')
                    if indici_perf[row[0].value][2] > 0:
                        ws[row[7].coordinate].font = Font(color='00B050')
                    else:
                        ws[row[7].coordinate].font = Font(color='FF0000')
                    ws[row[7].coordinate].alignment = Alignment(horizontal='center', vertical='center')
                    ws[row[7].coordinate].border = Border(bottom=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'))
                    ws[row[8].coordinate].value = "{0:.2f}%".format(round(indici_perf[row[0].value][3] * 100, 2)).replace('.', ',')
                    if indici_perf[row[0].value][3] > 0:
                        ws[row[8].coordinate].font = Font(color='00B050')
                    else:
                        ws[row[8].coordinate].font = Font(color='FF0000')
                    ws[row[8].coordinate].alignment = Alignment(horizontal='center', vertical='center')
                    ws[row[8].coordinate].border = Border(bottom=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'))
                    ws.merge_cells(start_row=row[5].row, end_row=row[5].row, start_column=row[5].column, end_column=row[6].column)
        
        # Textbox
        self.__textbox(ws, 6, 40, 10, 15)

        # Logo
        self.__logo(ws, col=6, colOff=0.8, row=43, rowOff=-0.2)

    def analisi_indici_5(self):
        """
        Crea la quinta pagina.
        Aggiunge fogli Indici_giornalieri.
        """
        # Carica indici giornalieri
        indici_giornalieri = pd.read_excel(
            self.file_portafoglio, 
            sheet_name='Indici_giornalieri', 
            names=['Date', 'S&P 500', 'Date.1', 'USDEUR', 'Date.2', 'VIX', 'Date.3', 'EURO STOXX 50']
        )
        indici_giornalieri['Date'] = pd.to_datetime(indici_giornalieri['Date'], format = '%Y-%m-%d %H:%M:%S').dt.strftime('%m-%Y')
        indici_giornalieri['Date.1'] = pd.to_datetime(indici_giornalieri['Date.1'], format = '%Y-%m-%d %H:%M:%S').dt.strftime('%m-%Y')
        indici_giornalieri['Date.2'] = pd.to_datetime(indici_giornalieri['Date.2'], format = '%Y-%m-%d %H:%M:%S').dt.strftime('%m-%Y')
        indici_giornalieri['Date.3'] = pd.to_datetime(indici_giornalieri['Date.3'], format = '%Y-%m-%d %H:%M:%S').dt.strftime('%m-%Y')
        # print(indici_giornalieri)
        
        # Aggiungi foglio dati per creare i grafici
        ws_dati_indici = self.wb.create_sheet('Dati_indici')
        ws_dati_indici = self.wb['Dati_indici']
        self.wb.active = ws_dati_indici
        # Carica gli indici giornalieri
        for r in dataframe_to_rows(indici_giornalieri, index=True, header=True):
            ws_dati_indici.append(r)
        ws_dati_indici.delete_rows(2)
        ws_dati_indici.sheet_state = 'hidden'

        ws = self.wb.create_sheet('5.an_mkt_perf')
        ws = self.wb['5.an_mkt_perf']
        self.wb.active = ws
        ws['A1'] = '1. Analisi Di Mercato'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells('A1:L4')

        # Aggiunta primo grafico
        chart = AreaChart()
        chart.title = "S&P 500"
        cats = Reference(ws_dati_indici, min_col=2, min_row=2, max_row=ws_dati_indici.max_row)
        data = Reference(ws_dati_indici, min_col=3, min_row=2, max_row=ws_dati_indici.max_row)
        chart.add_data(data, titles_from_data=False)
        series = chart.series[0]
        fill = PatternFillProperties()
        fill.foreground = ColorChoice(srgbClr='31869B')
        fill.background = ColorChoice(srgbClr='31869B')
        series.graphicalProperties.pattFill = fill
        chart.set_categories(cats)
        chart.height = 6.7
        chart.width = 10
        chart.legend = None
        ws.add_chart(chart, "A6")

        # Aggiunta secondo grafico
        chart = AreaChart()
        chart.title = "Usd/Eur"
        cats = Reference(ws_dati_indici, min_col=4, min_row=2, max_row=ws_dati_indici.max_row)
        data = Reference(ws_dati_indici, min_col=5, min_row=2, max_row=ws_dati_indici.max_row)
        chart.add_data(data, titles_from_data=False)
        series = chart.series[0]
        fill = PatternFillProperties()
        fill.foreground = ColorChoice(srgbClr='31869B')
        fill.background = ColorChoice(srgbClr='31869B')
        series.graphicalProperties.pattFill = fill
        chart.set_categories(cats)
        chart.height = 6.7
        chart.width = 10
        chart.legend = None
        ws.add_chart(chart, "G6")

        # Aggiunta terzo grafico
        chart = AreaChart()
        chart.title = "Vix"
        cats = Reference(ws_dati_indici, min_col=6, min_row=2, max_row=ws_dati_indici.max_row)
        data = Reference(ws_dati_indici, min_col=7, min_row=2, max_row=ws_dati_indici.max_row)
        chart.add_data(data, titles_from_data=False)
        series = chart.series[0]
        fill = PatternFillProperties()
        fill.foreground = ColorChoice(srgbClr='31869B')
        fill.background = ColorChoice(srgbClr='31869B')
        series.graphicalProperties.pattFill = fill
        chart.set_categories(cats)
        chart.height = 6.7
        chart.width = 10
        chart.legend = None
        ws.add_chart(chart, "A20")

        # Aggiunta quarto grafico
        chart = AreaChart()
        chart.title = "Eurostoxx 50"
        cats = Reference(ws_dati_indici, min_col=8, min_row=2, max_row=ws_dati_indici.max_row)
        data = Reference(ws_dati_indici, min_col=9, min_row=2, max_row=ws_dati_indici.max_row)
        chart.add_data(data, titles_from_data=False)
        series = chart.series[0]
        fill = PatternFillProperties()
        fill.foreground = ColorChoice(srgbClr='31869B')
        fill.background = ColorChoice(srgbClr='31869B')
        series.graphicalProperties.pattFill = fill
        chart.set_categories(cats)
        chart.height = 6.7
        chart.width = 10
        chart.legend = None
        ws.add_chart(chart, "G20")
        
        # Logo
        self.__logo(ws)

    def performance_6(self):
        """
        Crea la sesta pagina.
        """
        # 6.Performance
        ws = self.wb.create_sheet('6.perf')
        ws =  self.wb['6.perf']
        self.wb.active = ws
        # Corpo
        ws['A11'] = '2. Performance'
        ws['A11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A11'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A11'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells('A11:L14')

        # Logo
        self.__logo(ws)

    def andamento_7(self):
        """
        Crea la settima pagina.
        """
        # Carica performance benchmark
        benchmark = pd.read_excel(self.file_portafoglio, sheet_name='Benchmark', index_col=0, header=0)
        perf_bk_2007 = (float(benchmark.loc[self.t1, 'benchmark_2007']) - 100) / 100
        perf_bk_ytd = (float(benchmark.loc[self.t1, 'benchmark_2007']) - float(benchmark.loc[self.t0_ytd, 'benchmark_2007'])) / float(benchmark.loc[self.t0_ytd, 'benchmark_2007'])
        perf_month = (float(benchmark.loc[self.t1, 'benchmark_2007']) - float(benchmark.loc[self.t0_1m, 'benchmark_2007'])) / float(benchmark.loc[self.t0_1m, 'benchmark_2007'])
        # Carica portafoglio
        ptf = pd.read_excel(self.file_portafoglio, sheet_name='Portafoglio', index_col=0, header=0)
        # print(ptf['31/10/2015':'31/12/2021'])
        perf_ptf_2007 = (float(ptf.loc[self.t1, 'ptf_2007']) - 100) / 100
        perf_ptf_ytd = (float(ptf.loc[self.t1, 'ptf_2007']) - ptf.loc[self.t0_ytd, 'ptf_2007']) / ptf.loc[self.t0_ytd, 'ptf_2007']
        perf_ptf_month = (float(ptf.loc[self.t1, 'ptf_2007']) - ptf.loc[self.t0_1m, 'ptf_2007']) / ptf.loc[self.t0_1m, 'ptf_2007']


        ws = self.wb.create_sheet('7.andamento')
        ws = self.wb['7.andamento']
        self.wb.active = ws

        # Titolo
        ws['A1'] = 'Andamento Del Portafoglio'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells('A1:L4')

        # Aggiunta primo grafico
        chart = BarChart()
        chart.type = 'col'
        chart.title = "DA INIZIO MANDATO"
        chart.y_axis.scaling.min = min(perf_bk_2007, perf_ptf_2007, 0) - 0.04
        chart.y_axis.scaling.max = max(perf_bk_2007, perf_ptf_2007) + 0.04
        ws['A13'] = 'Ptf'
        ws['A14'] = perf_ptf_2007
        ws['A14'].number_format = FORMAT_PERCENTAGE_00
        ws['B13'] = 'Benchmark'
        ws['B14'] = perf_bk_2007
        ws['B14'].number_format = FORMAT_PERCENTAGE_00
        data = Reference(ws, min_col=1, max_col=2, min_row=13, max_row=14)
        chart.add_data(data, titles_from_data=True)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        s = chart.series[0]
        s.graphicalProperties.solidFill = 'FF6600'
        s1 = chart.series[1]
        s1.graphicalProperties.solidFill = '177245'
        chart.height = 10.3
        chart.width = 7.2
        chart.legend.position = 'b' # sposta la legenda in basso
        chart.y_axis.delete = True # togli l'asse y
        chart.x_axis.delete = True # togli l'asse x
        ws.add_chart(chart, 'A6')

        # Aggiunta secondo grafico
        chart = BarChart()
        chart.type = 'col'
        chart.title = "YEAR TO DATE"
        chart.y_axis.scaling.min = min(perf_bk_ytd, perf_ptf_ytd, 0) - 0.03
        chart.y_axis.scaling.max = max(perf_bk_ytd, perf_ptf_ytd, 0.04) + 0.03
        ws['F13'] = 'Ptf'
        ws['F14'] = perf_ptf_ytd
        ws['F14'].number_format = FORMAT_PERCENTAGE_00
        ws['G13'] = 'Benchmark'
        ws['G14'] = perf_bk_ytd
        ws['G14'].number_format = FORMAT_PERCENTAGE_00
        ws['H13'] = 'Target'
        ws['H14'] = 0.04
        ws['H14'].number_format = FORMAT_PERCENTAGE_00
        data = Reference(ws, min_col=6, max_col=8, min_row=13, max_row=14)
        chart.add_data(data, titles_from_data=True)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        s = chart.series[0]
        s.graphicalProperties.solidFill = 'FF6600'
        s1 = chart.series[1]
        s1.graphicalProperties.solidFill = '177245'
        s2 = chart.series[2]
        s2.graphicalProperties.solidFill = 'C00000'
        chart.legend.position = 'b' # sposta la legenda in basso
        chart.y_axis.delete = True # togli l'asse y
        chart.x_axis.delete = True # togli l'asse x
        size = XDRPositiveSize2D(pixels_to_EMU(272.12598), pixels_to_EMU(389.29))
        maker = AnchorMarker(col=4, colOff=0, row=5, rowOff=0)
        ancoraggio = OneCellAnchor(_from=maker, ext=size)
        ws.add_chart(chart)
        chart.anchor = ancoraggio

        # Aggiunta terzo grafico
        chart = BarChart()
        chart.type = 'col'
        chart.title = "MENSILE"
        chart.y_axis.scaling.min = -0.05
        chart.y_axis.scaling.max = 0.05
        ws['J13'] = 'Ptf'
        ws['J14'] = perf_ptf_month
        ws['J14'].number_format = FORMAT_PERCENTAGE_00
        ws['K13'] = 'Benchmark'
        ws['K14'] = perf_month
        ws['K14'].number_format = FORMAT_PERCENTAGE_00
        data = Reference(ws, min_col=10, max_col=11, min_row=13, max_row=14)
        chart.add_data(data, titles_from_data=True)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        s = chart.series[0]
        s.graphicalProperties.solidFill = 'FF6600'
        s1 = chart.series[1]
        s1.graphicalProperties.solidFill = '177245'
        chart.legend.position = 'b' # sposta la legenda in basso
        chart.y_axis.delete = True # togli l'asse y
        chart.x_axis.delete = True # togli l'asse x
        size = XDRPositiveSize2D(pixels_to_EMU(272.12598), pixels_to_EMU(389.29))
        maker = AnchorMarker(col=8, colOff=0, row=5, rowOff=0)
        ancoraggio = OneCellAnchor(_from=maker, ext=size)
        ws.add_chart(chart)
        chart.anchor = ancoraggio

        # Corpo
        ws['B27'] = '* il rendimento del P. in Strumenti è al netto della commissione di consulenza, degli eventuali prelievi e conferimenti'
        ws['B27'].font = Font(name='Times New Roman', size=11, bold=False, color='31869B')

        # Logo
        self.__logo(ws)

    def caricamento_dati(self):
        """
        Aggiunta fogli Cono, Portafoglio e Benchmark, poi nascosti.
        """
        # Carica dati per i coni
        coni = pd.read_excel(self.file_portafoglio, sheet_name='Cono', index_col=0, header=0)
        coni.index = pd.to_datetime(coni.index, format='%Y-%m-%d %H:%M:%S').strftime('%m-%Y')
        # Carica gli scenari per i coni
        ws_dati_cono = self.wb.create_sheet('Dati_cono')
        ws_dati_cono = self.wb['Dati_cono']
        self.wb.active = ws_dati_cono
        for r in dataframe_to_rows(coni, index=True, header=True):
            ws_dati_cono.append(r)
        ws_dati_cono.delete_rows(2)
        ws_dati_cono.sheet_state = 'hidden'

        # Carica rendimento portafoglio
        perf_pf = pd.read_excel(self.file_portafoglio, sheet_name='Portafoglio', index_col=0, header=0)
        perf_pf.index = pd.to_datetime(perf_pf.index, format='%Y-%m-%d %H:%M:%S').strftime('%m-%Y')
        # Carica perf ptf
        ws_dati_pf = self.wb.create_sheet('Dati_pf')
        ws_dati_pf = self.wb['Dati_pf']
        self.wb.active = ws_dati_pf
        for r in dataframe_to_rows(perf_pf, index=True, header=True):
            ws_dati_pf.append(r)
        ws_dati_pf.delete_rows(2)
        ws_dati_pf.sheet_state = 'hidden'

        # Carica performance benchmark
        perf_bk = pd.read_excel(self.file_portafoglio, sheet_name='Benchmark', index_col=0, header=0)
        perf_bk.index = pd.to_datetime(perf_bk.index, format='%Y-%m-%d %H:%M:%S').strftime('%m-%Y')
        # Carica perf bk
        ws_dati_bk = self.wb.create_sheet('Dati_bk')
        ws_dati_bk = self.wb['Dati_bk']
        self.wb.active = ws_dati_bk
        for r in dataframe_to_rows(perf_bk, index=True, header=True):
            ws_dati_bk.append(r)
        ws_dati_bk.delete_rows(2)
        ws_dati_bk.sheet_state = 'hidden'

    def cono_8(self):
        """
        Crea l'ottava pagina.
        Riattiva fogli ws_dati_bk, ws_dati_cono e ws_dati_pf.
        """
        # Riattiva scenari coni
        ws_dati_cono = self.wb['Dati_cono']
        # Riattiva performance ptf
        ws_dati_pf = self.wb['Dati_pf']
        # Riattiva performance bk
        ws_dati_bk = self.wb['Dati_bk']

        ws = self.wb.create_sheet('8.cono_1')
        ws = self.wb['8.cono_1']
        self.wb.active = ws

        # Titolo
        ws['A1'] = 'Cono Delle Probabilità'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells('A1:L4')

        # Corpo
        ws['E6'] = 'Benchmark 2016'
        ws['E6'].alignment = Alignment(horizontal='center', vertical='center')
        ws['E6'].font = Font(name='Times New Roman', size=14, bold=True, color='31869B')
        ws.merge_cells('E6:H6')

        # Aggiunta grafico
        chart = LineChart()
        riga_mese_t1 = lambda x : next(x[row[0].coordinate].row for row in x.iter_rows(min_col=0, max_col=0) if x[row[0].coordinate].value == self.t1.strftime('%m-%Y'))
        ws_dati_cono_max_row = riga_mese_t1(ws_dati_cono)
        # for row in ws_dati_cono.iter_rows(min_col=0, max_col=0):
        #     if ws_dati_cono[row[0].coordinate].value == self.t1.strftime('%m-%Y'):
        #         print(f"La riga del mese t1 è : {ws_dati_cono[row[0].coordinate].row}.")
        #         ws_dati_cono_max_row = ws_dati_cono[row[0].coordinate].row
        data = Reference(ws_dati_cono, min_col=7, max_col=9, min_row=109, max_row=ws_dati_cono_max_row) # hard coding
        chart.add_data(data, titles_from_data='False')
        # for row in ws_dati_bk.iter_rows(min_col=0, max_col=0):
        #     if ws_dati_bk[row[0].coordinate].value == self.t1.strftime('%m-%Y'):
        #         print(f"La riga del mese t1 è : {ws_dati_bk[row[0].coordinate].row}.")
        #         ws_dati_bk_max_row = ws_dati_bk[row[0].coordinate].row
        ws_dati_bk_max_row = riga_mese_t1(ws_dati_bk)
        data = Reference(ws_dati_bk, min_col=26, min_row=111, max_row=ws_dati_bk_max_row) # hard coding
        chart.add_data(data, titles_from_data='False')
        # for row in ws_dati_pf.iter_rows(min_col=0, max_col=0):
        #     if ws_dati_pf[row[0].coordinate].value == self.t1.strftime('%m-%Y'):
        #         print(f"La riga del mese t1 è : {ws_dati_pf[row[0].coordinate].row}.")
        #         ws_dati_pf_max_row = ws_dati_pf[row[0].coordinate].row
        ws_dati_pf_max_row = riga_mese_t1(ws_dati_pf)
        data = Reference(ws_dati_pf, min_col=5, min_row=109, max_row=ws_dati_pf_max_row) # hard coding
        chart.add_data(data, titles_from_data='False')

        s0 = chart.series[0]
        s0.graphicalProperties.line.solidFill = '0000FF'
        s0.graphicalProperties.line.width = 12700
        s0.dLbls = DataLabelList()
        dl = DataLabel(dLblPos='t', idx=ws_dati_cono_max_row-110, numFmt='0.00', showVal=True)
        s0.dLbls.dLbl.append(dl)
        s1 = chart.series[1]
        s1.graphicalProperties.line.solidFill = 'FF00FF'
        s1.graphicalProperties.line.width = 12700
        s1.dLbls = DataLabelList()
        dl = DataLabel(dLblPos='t', idx=ws_dati_cono_max_row-110, numFmt='0.00', showVal=True)
        s1.dLbls.dLbl.append(dl)
        s2 = chart.series[2]
        s2.graphicalProperties.line.solidFill = '000080'
        s2.graphicalProperties.line.width = 12700
        s2.dLbls = DataLabelList()
        dl = DataLabel(dLblPos='t', idx=ws_dati_cono_max_row-110, numFmt='0.00', showVal=True)
        s2.dLbls.dLbl.append(dl)
        s3 = chart.series[3]
        s3.graphicalProperties.line.solidFill = '177245'
        s3.graphicalProperties.line.width = 25400
        s3.dLbls = DataLabelList()
        dl = DataLabel(dLblPos='b', idx=ws_dati_bk_max_row-112, numFmt='0.00', showVal=True)
        s3.dLbls.dLbl.append(dl)
        s4 = chart.series[4]
        s4.graphicalProperties.line.solidFill = 'FF0000'
        s4.graphicalProperties.line.width = 25400
        s4.dLbls = DataLabelList()
        dl = DataLabel(dLblPos='t', idx=ws_dati_pf_max_row-110, numFmt='0.00', showVal=True)
        s4.dLbls.dLbl.append(dl)

        dates = Reference(ws_dati_cono, min_col=1, max_col=1, min_row=110, max_row=ws_dati_cono_max_row)
        chart.set_categories(dates)
        chart.legend.layout = Layout(manualLayout=ManualLayout(h=1))
        size = XDRPositiveSize2D(pixels_to_EMU(812.598), pixels_to_EMU(453.54))
        cellw = lambda x: cm_to_EMU((x * (18.65-1.71))/10)
        coloffset2 = cellw(0.1)
        maker = AnchorMarker(col=0, colOff=coloffset2, row=6, rowOff=0)
        ancoraggio = OneCellAnchor(_from=maker, ext=size)
        ws.add_chart(chart)
        chart.anchor = ancoraggio
        chart.y_axis.scaling.min = 95 # valore minimo asse y
        ws.row_dimensions[5].height = 11.25

        # Logo
        self.__logo(ws)

    def cono_9(self):
        """
        Crea la nona pagina.
        Riattiva fogli ws_dati_bk, ws_dati_cono e ws_dati_pf.
        """
        # Riattiva scenari coni
        ws_dati_cono = self.wb['Dati_cono']
        # Riattiva performance ptf
        ws_dati_pf = self.wb['Dati_pf']
        # Riattiva performance bk
        ws_dati_bk = self.wb['Dati_bk']

        ws = self.wb.create_sheet('9.cono_2')
        ws = self.wb['9.cono_2']
        self.wb.active = ws

        # Titolo
        ws['A1'] = 'Cono Delle Probabilità'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells('A1:L4')

        # Corpo
        ws['E6'] = 'Benchmark 2022'
        ws['E6'].alignment = Alignment(horizontal='center', vertical='center')
        ws['E6'].font = Font(name='Times New Roman', size=14, bold=True, color='31869B')
        ws.merge_cells('E6:H6')

        # Aggiunta grafico
        chart = LineChart()
        riga_mese_t1 = lambda x : next(x[row[0].coordinate].row for row in x.iter_rows(min_col=0, max_col=0) if x[row[0].coordinate].value == self.t1.strftime('%m-%Y'))
        ws_dati_cono_max_row = riga_mese_t1(ws_dati_cono)
        # for row in ws_dati_cono.iter_rows(min_col=0, max_col=0):
        #     if ws_dati_cono[row[0].coordinate].value == self.t1.strftime('%m-%Y'):
        #         print(f"La riga del mese t1 è : {ws_dati_cono[row[0].coordinate].row}.")
        #         ws_dati_cono_max_row = ws_dati_cono[row[0].coordinate].row
        data = Reference(ws_dati_cono, min_col=11, max_col=13, min_row=180, max_row=ws_dati_cono_max_row) # hard coding
        chart.add_data(data, titles_from_data='False')
        # for row in ws_dati_bk.iter_rows(min_col=0, max_col=0):
        #     if ws_dati_bk[row[0].coordinate].value == self.t1.strftime('%m-%Y'):
        #         print(f"La riga del mese t1 è : {ws_dati_bk[row[0].coordinate].row}.")
        #         ws_dati_bk_max_row = ws_dati_bk[row[0].coordinate].row
        ws_dati_bk_max_row = riga_mese_t1(ws_dati_bk)
        data = Reference(ws_dati_bk, min_col=27, min_row=182, max_row=ws_dati_bk_max_row) # hard coding
        chart.add_data(data, titles_from_data='False')
        # for row in ws_dati_pf.iter_rows(min_col=0, max_col=0):
        #     if ws_dati_pf[row[0].coordinate].value == self.t1.strftime('%m-%Y'):
        #         print(f"La riga del mese t1 è : {ws_dati_pf[row[0].coordinate].row}.")
        #         ws_dati_pf_max_row = ws_dati_pf[row[0].coordinate].row
        ws_dati_pf_max_row = riga_mese_t1(ws_dati_pf)
        data = Reference(ws_dati_pf, min_col=6, min_row=180, max_row=ws_dati_pf_max_row) # hard coding
        chart.add_data(data, titles_from_data='False')

        s0 = chart.series[0]
        s0.graphicalProperties.line.solidFill = '0000FF'
        s0.graphicalProperties.line.width = 12700
        s0.dLbls = DataLabelList()
        dl = DataLabel(dLblPos='t', idx=ws_dati_cono_max_row-181, numFmt='0.00', showVal=True)
        s0.dLbls.dLbl.append(dl)
        s1 = chart.series[1]
        s1.graphicalProperties.line.solidFill = 'FF00FF'
        s1.graphicalProperties.line.width = 12700
        s1.dLbls = DataLabelList()
        dl = DataLabel(dLblPos='t', idx=ws_dati_cono_max_row-181, numFmt='0.00', showVal=True)
        s1.dLbls.dLbl.append(dl)
        s2 = chart.series[2]
        s2.graphicalProperties.line.solidFill = '000080'
        s2.graphicalProperties.line.width = 12700
        s2.dLbls = DataLabelList()
        dl = DataLabel(dLblPos='t', idx=ws_dati_cono_max_row-181, numFmt='0.00', showVal=True)
        s2.dLbls.dLbl.append(dl)
        s3 = chart.series[3]
        s3.graphicalProperties.line.solidFill = '177245'
        s3.graphicalProperties.line.width = 25400
        s3.dLbls = DataLabelList()
        dl = DataLabel(dLblPos='b', idx=ws_dati_bk_max_row-183, numFmt='0.00', showVal=True)
        s3.dLbls.dLbl.append(dl)
        s4 = chart.series[4]
        s4.graphicalProperties.line.solidFill = 'FF0000'
        s4.graphicalProperties.line.width = 25400
        s4.dLbls = DataLabelList()
        dl = DataLabel(dLblPos='t', idx=ws_dati_pf_max_row-181, numFmt='0.00', showVal=True)
        s4.dLbls.dLbl.append(dl)

        dates = Reference(ws_dati_cono, min_col=1, max_col=1, min_row=181, max_row=ws_dati_cono_max_row)
        chart.set_categories(dates)
        chart.legend.layout = Layout(manualLayout=ManualLayout(h=1))
        size = XDRPositiveSize2D(pixels_to_EMU(812.598), pixels_to_EMU(453.54))
        cellw = lambda x: cm_to_EMU((x * (18.65-1.71))/10)
        coloffset2 = cellw(0.1)
        maker = AnchorMarker(col=0, colOff=coloffset2, row=6, rowOff=0)
        ancoraggio = OneCellAnchor(_from=maker, ext=size)
        ws.add_chart(chart)
        chart.anchor = ancoraggio
        chart.y_axis.scaling.min = 90 # valore minimo asse y
        ws.row_dimensions[5].height = 11.25

        # Logo
        self.__logo(ws)

    def nuovo_bk_10(self):
        """
        Crea la decima pagina.
        """
        ws = self.wb.create_sheet('10.nuovo_bk')
        ws = self.wb['10.nuovo_bk']
        self.wb.active = ws
        # Titolo
        ws['A1'] = 'Nuovo Benchmark'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells('A1:L4')
        # Corpo
        body_10_1 = [
            'Indice MTS BOT', 'Bloomberg Euro Government', 'Bloomberg Euro Corporate Index', 'Bloomberg Pan-European High Yield Index',
            'Bloomberg Global Aggregate Index', 'MSCI Europa', 'MSCI USA', 'MSCI Pacifico', 'MSCI Emerging Market Free', 
            'HFRX Absolute Return', 'Bloomberg Commodity Index'
        ]
        body_10_2 = ['20,00%', '13,15%', '4,66%', '8,05%', '10,09%', '7,27%', '14,55%', '4,56%', '11,04%', '2,85%', '3,78%']
        for row in ws.iter_rows(min_row=8, max_row=8+len(body_10_1)-1, min_col=4, max_col=9):
            ws[row[0].coordinate].value = body_10_1[0]
            del body_10_1[0]
            ws[row[0].coordinate].font = Font(name='Calibri', size=11, bold=True, italic=True, color='000000') 
            ws[row[0].coordinate].border = Border(bottom=Side(border_style='mediumDashDot', color='31869B'))
            ws.merge_cells(start_row=row[0].row, end_row=row[0].row, start_column=row[0].column, end_column=row[4].column)
            ws[row[5].coordinate].value = body_10_2[0]
            del body_10_2[0]
            ws[row[5].coordinate].font = Font(name='Calibri', size=11, bold=True, italic=True, color='000000') 
            ws[row[5].coordinate].alignment = Alignment(horizontal='right')
            ws[row[5].coordinate].border = Border(bottom=Side(border_style='mediumDashDot', color='31869B'))
        ws['C21'] = '           Benchmark costruito seguendo la composizione del portafoglio al 31/12/2021'
        ws['C21'].font = Font(name='Calibri', size=11, bold=True, italic=True, color='31869B') 
        # Logo
        self.__logo(ws)

    def performance_11(self):
        """
        Crea l'undicesima pagina.
        """
        # Carica portafoglio
        portfolio = pd.read_excel(self.file_portafoglio, sheet_name='Portfolio', header=1)
        # Carica performance posizioni --- dipende da portfolio
        delta = pd.read_excel(self.file_portafoglio, sheet_name='Delta', index_col=0, header=0)
        # print(delta)
        ws = self.wb.create_sheet('11.perf_mese')
        ws = self.wb['11.perf_mese']
        self.wb.active = ws
        min_row = 1
        min_col = 1
        # Titolo
        ws['A1'] = 'Performance Del Mese'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells('A1:L4')
        min_row += 5
        # Colonne
        header_11 = ['', '', 'Totale ' + self.mesi_dict[self.t0_1m.month], 'Totale ' + self.mesi_dict[self.t1.month], 'Δ', 'Δ%', 'Δ% YTD']
        len_header_11 = len(header_11)
        for column in ws.iter_cols(min_col=min_col, max_col=min_col+len_header_11-1, min_row=min_row, max_row=min_row):
            ws[column[0].coordinate].value = header_11[0]
            del header_11[0]
            ws[column[0].coordinate].font = Font(name='Times New Roman', size=10, color='FFFFFF')
            ws[column[0].coordinate].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[column[0].coordinate].fill = PatternFill(fill_type='solid', fgColor='31869B')
        # Indice
        intermediari = portfolio.loc[:, 'INTERMEDIARIO'].unique()
        intermediari = list(intermediari)
        intermediari.insert(len(intermediari), 'Interessi Phoenix')
        intermediari.insert(len(intermediari), 'Totale Complessivo')
        # Rimuovi gli intermediari per cui non calcolo la performance
        intermediari.remove('Banca Valsabbina Nespoli')
        intermediari.remove('Crédit Agricole Artes')
        intermediari.remove('Crédit Agricole B.N.')
        intermediari.remove('Altro')
        len_int = len(intermediari)
        # Corpo tabella
        for row in ws.iter_rows(min_col=min_col, max_col=min_col+len_header_11-1, min_row=min_row+1, max_row=min_row+1+len_int-1):
            # if intermediari[0] == 'Mediolanum':
            #     intermediari = np.delete(intermediari, 0)
            #     continue
            ws[row[0].coordinate].value = intermediari[0]
            del intermediari[0]
            ws.column_dimensions[row[2].column_letter].width = 10.5
            ws.column_dimensions[row[3].column_letter].width = 10.5
            ws.row_dimensions[row[0].row].height = 25.50
            ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, color='000000')
            ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws[row[0].coordinate].border = Border(left=Side(border_style='dashed', color='31869B'), bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'))
            ws.merge_cells(start_column=row[0].column, end_column=row[1].column, start_row=row[0].row, end_row=row[0].row)
            # Corpo tabella
            ws[row[2].coordinate].border = Border(right=Side(border_style='dashed', color='31869B'), bottom=Side(border_style='dashed', color='31869B'))
            ws[row[2].coordinate].number_format = '€ #,0'
            ws[row[2].coordinate].font = Font(name='Times New Roman', size=9)
            ws[row[2].coordinate].alignment = Alignment(horizontal='center', vertical='center')
            ws[row[3].coordinate].border = Border(right=Side(border_style='dashed', color='31869B'), bottom=Side(border_style='dashed', color='31869B'))
            ws[row[3].coordinate].number_format = '€ #,0'
            ws[row[3].coordinate].font = Font(name='Times New Roman', size=9)
            ws[row[3].coordinate].alignment = Alignment(horizontal='center', vertical='center')
            ws[row[4].coordinate].border = Border(right=Side(border_style='dashed', color='31869B'), bottom=Side(border_style='dashed', color='31869B'))
            ws[row[4].coordinate].number_format = '€ #,0'
            ws[row[4].coordinate].font = Font(name='Times New Roman', size=9)
            ws[row[4].coordinate].alignment = Alignment(horizontal='center', vertical='center')
            ws[row[5].coordinate].border = Border(right=Side(border_style='dashed', color='31869B'), bottom=Side(border_style='dashed', color='31869B'))
            ws[row[5].coordinate].number_format = FORMAT_PERCENTAGE_00
            ws[row[5].coordinate].font = Font(name='Times New Roman', size=9)
            ws[row[5].coordinate].alignment = Alignment(horizontal='center', vertical='center')
            ws[row[6].coordinate].border = Border(right=Side(border_style='dashed', color='31869B'), bottom=Side(border_style='dashed', color='31869B'))
            ws[row[6].coordinate].number_format = FORMAT_PERCENTAGE_00
            ws[row[6].coordinate].font = Font(name='Times New Roman', size=9)
            ws[row[6].coordinate].alignment = Alignment(horizontal='center', vertical='center')
            
            if ws[row[0].coordinate].row != 7+len_int-1:
                ws[row[2].coordinate].value = delta.loc[ws[row[0].coordinate].value, 'Totale mese passato']
                ws[row[3].coordinate].value = delta.loc[ws[row[0].coordinate].value, 'Totale mese corrente']
                ws[row[4].coordinate].value = delta.loc[ws[row[0].coordinate].value, 'Δ']
                ws[row[5].coordinate].value = delta.loc[ws[row[0].coordinate].value, 'Δ%']
                ws[row[6].coordinate].value = delta.loc[ws[row[0].coordinate].value, 'Δ% YTD']
            else:
                for _ in range(0, len_header_11):
                    ws[row[_].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
                    ws[row[_].coordinate].border = Border(top=Side(border_style='medium', color='31869B'), right=Side(border_style='medium', color='31869B'), left=Side(border_style='medium', color='31869B'), bottom=Side(border_style='medium', color='31869B'))
                    ws[row[4].coordinate].value = 0
                    for __ in range(1, len_int): # Somma per tutti i valori nella colonna delta
                        ws[row[4].coordinate].value = ws[row[4].coordinate].value + ws[row[4].coordinate].offset(row=-__).value

        # Textbox
        self.__textbox(ws, min_row, min_row + len_int, 8, 12)

        # Logo
        self.__logo(ws, colOff=0, row=min_row + len_int + 8)

    def tabella_prezzi(self, ws: Worksheet, min_row: int, intermediario: str, strumenti: pd.DataFrame) -> int:
        """Crea una tabella dei prezzi degli strumenti di un intermediario

        Arguments:
            ws {Worksheet} -- sheet
            min_row {int} -- minimum row
            intermediario {str} -- broker
            strumenti {pd.DataFrame} -- assets

        Returns:
            int -- updated minimum row
        """
        min_row = min_row
        min_col = 1
        max_col = 12
        # Titolo
        ws.cell(row=min_row, column=min_col, value=intermediario) 
        ws.cell(row=min_row, column=min_col).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=min_row, column=min_col).font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws.cell(row=min_row, column=min_col).fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells(start_row=min_row, start_column=min_col, end_row=min_row+3, end_column=max_col)
        min_row += 5
        # Creazione tabella
        header = ['Nome', '', '', '', '', 'Valuta', 'Quantità', 'Prezzo di carico', 'Prezzo attuale', '∆ prezzo', 'Ctv', '']
        for column in ws.iter_cols(min_col=min_col, max_col=min_col + len(header) - 1, min_row=min_row, max_row=min_row):
            ws[column[0].coordinate].value = header[0]
            del header[0]
            ws[column[0].coordinate].font = Font(name='Times New Roman', size=10, color='FFFFFF', bold=True)
            ws[column[0].coordinate].alignment = Alignment(horizontal='center', vertical='center')
            ws[column[0].coordinate].fill = PatternFill(fill_type='solid', fgColor='31869B')
        mid_col = min_col
        ws.merge_cells(start_row=min_row, start_column=mid_col, end_row=min_row+1, end_column=mid_col+4)
        mid_col += 5
        ws.merge_cells(start_row=min_row, start_column=mid_col, end_row=min_row+1, end_column=mid_col)
        mid_col += 1
        ws.merge_cells(start_row=min_row, start_column=mid_col, end_row=min_row+1, end_column=mid_col)
        mid_col += 1
        ws.merge_cells(start_row=min_row, start_column=mid_col, end_row=min_row+1, end_column=mid_col)
        ws.cell(row=min_row, column=mid_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        mid_col += 1
        ws.merge_cells(start_row=min_row, start_column=mid_col, end_row=min_row+1, end_column=mid_col)
        ws.cell(row=min_row, column=mid_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        mid_col += 1
        ws.merge_cells(start_row=min_row, start_column=mid_col, end_row=min_row+1, end_column=mid_col)
        mid_col += 1
        ws.merge_cells(start_row=min_row, start_column=mid_col, end_row=min_row+1, end_column=mid_col+1)
        min_row += 2
        # Corpo tabella
        prodotti = strumenti['PRODOTTO'].to_list()
        for row in ws.iter_rows(min_row=min_row, max_row=strumenti.shape[0] + min_row - 1, min_col=1, max_col=12):
            ws[row[0].coordinate].value = prodotti[0]
            del prodotti[0]
            ws[row[0].coordinate].font = Font(name='Times New Roman', size=10)
            ws[row[5].coordinate].value = (strumenti.loc[strumenti['PRODOTTO']==ws[row[0].coordinate].value, 'DIVISA']).values[0]
            ws[row[5].coordinate].font = Font(name='Times New Roman', size=10)
            ws[row[5].coordinate].alignment = Alignment(horizontal='center')
            ws[row[6].coordinate].value = (strumenti.loc[strumenti['PRODOTTO']==ws[row[0].coordinate].value, 'QUANTITA t1']).values[0]
            ws[row[6].coordinate].font = Font(name='Times New Roman', size=10)
            ws[row[6].coordinate].number_format = '#,##0.00'
            ws[row[7].coordinate].value = (strumenti.loc[strumenti['PRODOTTO']==ws[row[0].coordinate].value, 'prezzo_di_carico']).values[0]
            ws[row[7].coordinate].font = Font(name='Times New Roman', size=10)
            ws[row[7].coordinate].number_format = FORMAT_NUMBER_00
            ws[row[8].coordinate].value = (strumenti.loc[strumenti['PRODOTTO']==ws[row[0].coordinate].value, 'PREZZO t1']).values[0]
            ws[row[8].coordinate].font = Font(name='Times New Roman', size=10)
            ws[row[8].coordinate].number_format = FORMAT_NUMBER_00
            ws[row[9].coordinate].value = (ws[row[8].coordinate].value / ws[row[7].coordinate].value) - 1
            ws[row[9].coordinate].font = Font(name='Times New Roman', size=10)
            ws[row[9].coordinate].alignment = Alignment(horizontal='center')
            ws[row[9].coordinate].number_format = FORMAT_PERCENTAGE_00
            ws[row[10].coordinate].value = (strumenti.loc[strumenti['PRODOTTO']==ws[row[0].coordinate].value, 'TOTALE t1']).values[0]
            ws[row[10].coordinate].font = Font(name='Times New Roman', size=10)
            ws[row[10].coordinate].alignment = Alignment(horizontal='center')
            ws[row[10].coordinate].number_format = '€ #,##0.00'
            ws.merge_cells(start_row=row[0].row, end_row=row[0].row, start_column=row[0].column, end_column=row[4].column)
            ws.merge_cells(start_row=row[0].row, end_row=row[0].row, start_column=row[10].column, end_column=row[11].column)
        min_row += strumenti.shape[0]
        return min_row + 2

    def prezzi_12(self):
        """
        Crea la dodicesima pagina.
        Strumenti di Banca Patrimoni e Banca Valsabbina Artes
        """
        # Carica portafoglio
        portfolio = pd.read_excel(self.file_portafoglio, sheet_name='Portfolio', header=1)
        # Crea foglio
        ws = self.wb.create_sheet('12.prezzi')
        ws = self.wb['12.prezzi']
        self.wb.active = ws
        min_row = 1
        # Crea tabella Banca Patrimoni
        banca_patrimoni_strumenti_liquidi = portfolio[
            ((portfolio['INTERMEDIARIO']=='Banca Patrimoni Nespoli') | 
             (portfolio['INTERMEDIARIO']=='Banca Patrimoni Artes') |
             (portfolio['INTERMEDIARIO']=='Banca Patrimoni Trust')) &
               (~portfolio['CATEGORIA'].isin(['CASH', 'CASH_FOREIGN_CURR', 'ALTERNATIVE_ASSET', 'GP']))]
        min_row = self.tabella_prezzi(ws, min_row, 'Banca Patrimoni', banca_patrimoni_strumenti_liquidi)
        # Crea tabella Banca Valsabbina
        banca_valsabbina_strumenti_liquidi = portfolio[
            (portfolio['INTERMEDIARIO']=='Banca Valsabbina Artes') &
              (~portfolio['CATEGORIA'].isin(['CASH', 'CASH_FOREIGN_CURR', 'ALTERNATIVE_ASSET', 'GP']))]
        min_row = self.tabella_prezzi(ws, min_row, 'Banca Valsabbina', banca_valsabbina_strumenti_liquidi)
        # Logo
        if len(banca_patrimoni_strumenti_liquidi+banca_valsabbina_strumenti_liquidi) > 26:
            self.__logo(ws, row=35+(len(banca_patrimoni_strumenti_liquidi+banca_valsabbina_strumenti_liquidi)-26))
        else:
            self.__logo(ws)

    def prezzi_13(self):
        """
        Crea la tredicesima pagina.
        Strumenti di Corner.
        """
        # Carica portafoglio
        portfolio = pd.read_excel(self.file_portafoglio, sheet_name='Portfolio', header=1)
        # Crea foglio
        ws = self.wb.create_sheet('13.prezzi')
        ws = self.wb['13.prezzi']
        self.wb.active = ws
        min_row = 1
        # Crea tabella Corner
        corner_strumenti_liquidi = portfolio[
            (portfolio['INTERMEDIARIO']=='Corner') &
              (~portfolio['CATEGORIA'].isin(['CASH', 'CASH_FOREIGN_CURR', 'ALTERNATIVE_ASSET', 'GP']))]
        min_row = self.tabella_prezzi(ws, min_row, 'Corner', corner_strumenti_liquidi)
        # Logo
        if len(corner_strumenti_liquidi) > 26:
            self.__logo(ws, row=35+(len(corner_strumenti_liquidi)-26))
        else:
            self.__logo(ws)

    def prezzi_14(self):
        """
        Crea la quattordicesima pagina.
        Strumenti di Mediobanca e Mediolanum
        """
        # Carica portafoglio
        portfolio = pd.read_excel(self.file_portafoglio, sheet_name='Portfolio', header=1)
        # Crea foglio
        ws = self.wb.create_sheet('14.prezzi')
        ws = self.wb['14.prezzi']
        self.wb.active = ws
        min_row = 1
        # Crea tabella Mediobanca
        mediobanca_strumenti_liquidi = portfolio[
            (portfolio['INTERMEDIARIO']=='Mediobanca') &
              (~portfolio['CATEGORIA'].isin(['CASH', 'CASH_FOREIGN_CURR', 'ALTERNATIVE_ASSET', 'GP']))]
        min_row = self.tabella_prezzi(ws, min_row, 'Mediobanca', mediobanca_strumenti_liquidi)
        # Crea tabella Mediolanum
        mediolanum_strumenti_liquidi = portfolio[
            (portfolio['INTERMEDIARIO']=='Mediolanum') &
              (~portfolio['CATEGORIA'].isin(['CASH', 'CASH_FOREIGN_CURR', 'ALTERNATIVE_ASSET', 'GP']))]
        min_row = self.tabella_prezzi(ws, min_row, 'Mediolanum', mediolanum_strumenti_liquidi)
        # Logo
        if len(mediobanca_strumenti_liquidi+mediolanum_strumenti_liquidi) > 26:
            self.__logo(
                ws, row=35+(len(mediobanca_strumenti_liquidi+mediolanum_strumenti_liquidi)-26)
                )
        else:
            self.__logo(ws)

    def att_in_corso_15(self):
        """
        Crea la quindicesima pagina
        """
        ws = self.wb.create_sheet('15.att')
        ws = self.wb['15.att']
        self.wb.active = ws

        ws['A1'] = 'Attività Svolte Ed In Corso'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells('A1:L4')

        # Text box
        self.__textbox(ws, 6, 28, 1, 12)

        self.__logo(ws)

    def valutazione_per_macroclasse_16(self):
        """
        Crea la sedicesima pagina.
        """
        ws = self.wb.create_sheet('16.vpm')
        ws = self.wb['16.vpm']
        self.wb.active = ws

        ws['A11'] = '3. Valutazione Per Macroclasse'
        ws['A11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A11'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells('A11:L14')
        ws['A11'].font = Font(name='Times New Roman', size=36, bold=True, color='FFFFFF')

        self.__logo(ws)

    def sintesi_17(self):
        """
        Crea la diciasettesima pagina.
        """
        # Carica portafoglio
        portfolio = pd.read_excel(self.file_portafoglio, sheet_name='Portfolio', header=1)

        ws = self.wb.create_sheet('17.sintesi')
        ws = self.wb['17.sintesi']
        self.wb.active = ws

        # Creazione tabella
        #print(portfolio['INTERMEDIARIO'].unique())
        header_17 = list(portfolio['INTERMEDIARIO'].unique())
        header_17.insert(0, '')
        header_17.extend(('Totale '+ self.mesi_dict[self.t1.month], 'Totale '+ self.mesi_dict[self.t0_1m.month]))
        len_header_17 = len(header_17)
        #print(header_17)

        # Titolo
        ws['A1'] = 'Sintesi'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        if len(list(portfolio['INTERMEDIARIO'].unique())) == 1:
            lunghezza_titolo_17 = 12
            min_col = 4
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 2:
            lunghezza_titolo_17 = 12
            min_col = 4
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 3:
            lunghezza_titolo_17 = 12
            min_col = 3
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 4:
            lunghezza_titolo_17 = 12
            min_col = 3
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 5:
            lunghezza_titolo_17 = 12
            min_col = 2
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 6:
            lunghezza_titolo_17 = 12
            min_col = 2
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 7:
            lunghezza_titolo_17 = 12
            min_col = 1
        else:
            lunghezza_titolo_17 = len_header_17
            min_col = 1
        ws.merge_cells(start_row=1, end_row=4, start_column=1, end_column=lunghezza_titolo_17)

        for col in ws.iter_cols(min_row=8, max_row=9, min_col=min_col, max_col=min_col + len_header_17 - 1):
            ws[col[0].coordinate].value = header_17[0]
            del header_17[0]
            ws[col[0].coordinate].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col[0].coordinate].font = Font(name='Times New Roman', size=10, color='FFFFFF')
            ws[col[0].coordinate].fill = PatternFill(fill_type='solid', fgColor='31869B')
            ws[col[0].coordinate].border = Border(right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'))
            ws.merge_cells(start_row=col[0].row, end_row=col[1].row, start_column=col[0].column, end_column=col[0].column)
            ws.row_dimensions[col[0].row].height = 20
            ws.row_dimensions[col[1].row].height = 20
            ws.column_dimensions[col[0].column_letter].width = 12

        tipo_strumento = list(portfolio['CATEGORIA'].unique())
        len_tipo_strumento = len(tipo_strumento)
        num_intermediari = len(portfolio['INTERMEDIARIO'].unique())
        #print(num_intermediari)
        lunghezza_colonna_17 = []
        # print(tipo_strumento)
        # print(type(tipo_strumento))
        # print(len(tipo_strumento))
        tipo_strumento_dict = {
            'CASH' : 'LIQUIDITÀ', 'GP' : 'GESTIONI', 'EQUITY' : 'AZIONI', 'CASH_FOREIGN_CURR' : 'LIQUIDITÀ IN VALUTA', 
            'CORPORATE_BOND' : 'OBBLIGAZIONI CORPORATE', 'GOVERNMENT_BOND' : 'OBBLIGAZIONI GOVERNATIVE', 
            'ALTERNATIVE_ASSET' : 'INVESTIMENTI ALTERNATIVI', 'HEDGE_FUND' : 'HEDGE FUND'
        }
        #tipo_strumento_dict = {k: v for k, v in sorted(tipo_strumento_dict.items(), key=lambda item: item[1])}

        for row in ws.iter_rows(min_row=8, max_row=10 + len_tipo_strumento -1, min_col=min_col, max_col=min_col + len_header_17):
            if row[0].row > 9:
                #print(ws[row[0].coordinate])
                #ws[row[0].coordinate].value = tipo_strumento_dict[tipo_strumento[0]]
                ws[row[0].coordinate].value = tipo_strumento[0] # carica i tipi di strumenti nell'indice
                del tipo_strumento[0]
                ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
                ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, color='000000')
                ws[row[0].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws.row_dimensions[row[0].row].height = 19
                #print(ws[row[1].offset(column=0, row=-2-(row[0].row-10)).coordinate].value) # mostra sempre l'intermediario
                for _ in range(1, num_intermediari+1):
                    ws[row[_].coordinate].value = portfolio.loc[(portfolio['INTERMEDIARIO']==ws[row[_].offset(column=0, row=-2-(row[0].row-10)).coordinate].value) & (portfolio['CATEGORIA']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() if portfolio.loc[(portfolio['INTERMEDIARIO']==ws[row[_].offset(column=0, row=-2-(row[0].row-10)).coordinate].value) & (portfolio['CATEGORIA']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() != 0 else ''
                    ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                    ws[row[_].coordinate].font = Font(name='Times New Roman', size=9)
                    ws[row[_].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                    ws[row[_].coordinate].number_format = '#,0'
                    # somma=[]
                    # somma.append(ws[row[_].coordinate].value)
                    # print(somma)
                    #ws[row[num_intermediari+1].coordinate].value = 0
                    #ws[row[num_intermediari+1].coordinate].value = (ws[row[num_intermediari+1].coordinate].value + float(ws[row[_].coordinate].value)) if float(ws[row[_].coordinate].value) != ''
                    #ws.cell(row=row[_].row, column=num_intermediari+1, value=ws[row[_].coordinate].value) += 
                    #print(ws[row[_].coordinate].value)
                    # print(ws[row[num_intermediari+1].coordinate].value)
                    # ws[row[num_intermediari+1].coordinate].value = 0
                    # try:
                    #     ws[row[num_intermediari+1].coordinate].value += float(ws[row[_].coordinate].value)
                    # except ValueError:
                    #     pass
                    # except TypeError:
                    #     pass
                    # if float(ws[row[_].coordinate].value) != ''
                #ws[row[num_intermediari+2].coordinate].value = '=SUM('+str(ws[row[1].coordinate])+':'+str(ws[row[num_intermediari].coordinate])+')'
                # Somma per strumenti
                ws[row[num_intermediari+1].coordinate].value = portfolio.loc[portfolio['CATEGORIA']==ws[row[0].coordinate].value, 'TOTALE t1'].sum()
                ws[row[num_intermediari+1].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+1].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+1].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+1].coordinate].number_format = '#,0'
                ws[row[num_intermediari+2].coordinate].value = portfolio.loc[portfolio['CATEGORIA']==ws[row[0].coordinate].value, 'TOTALE t0'].sum()
                ws[row[num_intermediari+2].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+2].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+2].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+2].coordinate].number_format = '#,0'

                ws[row[0].coordinate].value = tipo_strumento_dict[ws[row[0].coordinate].value] # aggiorna valori dell'indice con i nomi nel dizionario
                lunghezza_colonna_17.append(len(ws.cell(row=row[0].row, column=row[0].column).value)) # ottieni la lunghezza della colonna
                ws.column_dimensions[row[0].column_letter].width = max(lunghezza_colonna_17) + 2.5 # modifica larghezza colonna

        # Somma per intermediari
        for row in ws.iter_rows(min_row=10 + len_tipo_strumento, max_row=10 + len_tipo_strumento, min_col=min_col, max_col=min_col + len_header_17):
            ws[row[0].coordinate].value = 'TOTALE'
            ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
            ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
            ws[row[0].coordinate].border = Border(bottom=Side(border_style='thin', color='31869B'), right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'), top=Side(border_style='thin', color='31869B'))
            ws.row_dimensions[row[0].row].height = 19
            #print(ws.cell(row=row[1].row, column=row[1].column).offset(row=-len_tipo_strumento))
            for _ in range(1,len_header_17-2):
                ws[row[_].coordinate].value = portfolio.loc[portfolio['INTERMEDIARIO']==ws.cell(row=row[_].row, column=row[_].column).offset(row=-len_tipo_strumento-2).value, 'TOTALE t1'].sum()
                # assert ws[row[_].coordinate].value == 'SUM(B10:B17)'
            ws[row[len_header_17-2].coordinate].value = portfolio.loc[:, 'TOTALE t1'].sum()
            ws[row[len_header_17-1].coordinate].value = portfolio.loc[:, 'TOTALE t0'].sum()
            for _ in range(1,len_header_17):
                ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                ws[row[_].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
                ws[row[_].coordinate].border = Border(bottom=Side(border_style='thin', color='31869B'), right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'), top=Side(border_style='thin', color='31869B'))
                ws[row[_].coordinate].number_format = '#,0'

    def valuta_18(self):
        """
        Crea la diciottesima pagina.
        """
        # Carica portafoglio
        portfolio = pd.read_excel(self.file_portafoglio, sheet_name='Portfolio', header=1)

        ws = self.wb.create_sheet('18.valuta')
        ws = self.wb['18.valuta']
        self.wb.active = ws

        # Creazione tabella
        header_18 = list(portfolio['INTERMEDIARIO'].unique())
        header_18.insert(0, '')
        header_18.extend(('Totale '+ self.mesi_dict[self.t1.month], 'Totale '+ self.mesi_dict[self.t0_1m.month]))
        len_header_18 = len(header_18)

        # Titolo
        ws['A1'] = 'Valuta'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        if len(list(portfolio['INTERMEDIARIO'].unique())) == 1:
            lunghezza_titolo_18 = 12
            min_col = 4
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 2:
            lunghezza_titolo_18 = 12
            min_col = 4
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 3:
            lunghezza_titolo_18 = 12
            min_col = 3
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 4:
            lunghezza_titolo_18 = 12
            min_col = 3
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 5:
            lunghezza_titolo_18 = 12
            min_col = 2
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 6:
            lunghezza_titolo_18 = 12
            min_col = 2
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 7:
            lunghezza_titolo_18 = 12
            min_col = 1
        else:
            lunghezza_titolo_18 = len_header_18
            min_col = 1
        ws.merge_cells(start_row=1, end_row=4, start_column=1, end_column=lunghezza_titolo_18)

        for col in ws.iter_cols(min_row=8, max_row=9, min_col=min_col, max_col=min_col + len_header_18 - 1):
            ws[col[0].coordinate].value = header_18[0]
            del header_18[0]
            ws[col[0].coordinate].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col[0].coordinate].font = Font(name='Times New Roman', size=10, color='FFFFFF')
            ws[col[0].coordinate].fill = PatternFill(fill_type='solid', fgColor='31869B')
            ws[col[0].coordinate].border = Border(right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'))
            ws.merge_cells(start_row=col[0].row, end_row=col[1].row, start_column=col[0].column, end_column=col[0].column)
            ws.row_dimensions[col[0].row].height = 20
            ws.row_dimensions[col[1].row].height = 20
            ws.column_dimensions[col[0].column_letter].width = 12

        tipo_divisa = list(portfolio['DIVISA'].unique())
        tipo_divisa.sort()
        # tipo_divisa.insert(len(tipo_divisa), 'ALTRE VALUTE')
        len_tipo_divisa = len(tipo_divisa)
        num_intermediari = len(portfolio['INTERMEDIARIO'].unique())
        lunghezza_colonna_18 = []
        #tipo_divisa_dict = {'CASH' : 'LIQUIDITÀ', 'GP' : 'GESTIONI', 'EQUITY' : 'AZIONI', 'CASH_FOREIGN_CURR' : 'LIQUIDITÀ IN VALUTA', 'CORPORATE_BOND' : 'OBBLIGAZIONI CORPORATE', 'GOVERNMENT_BOND' : 'OBBLIGAZIONI GOVERNATIVE', 'ALTERNATIVE_ASSET' : 'INVESTIMENTI ALTERNATIVI', 'HEDGE_FUND' : 'HEDGE FUND'}
        for row in ws.iter_rows(min_row=8, max_row=10 + len_tipo_divisa -1, min_col=min_col, max_col=min_col + len_header_18):
            if row[0].row > 9:
                ws[row[0].coordinate].value = tipo_divisa[0]
                del tipo_divisa[0]
                ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
                ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, color='000000')
                ws[row[0].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws.row_dimensions[row[0].row].height = 19
                for _ in range(1, num_intermediari+1):
                    ws[row[_].coordinate].value = portfolio.loc[(portfolio['INTERMEDIARIO']==ws[row[_].offset(column=0, row=-2-(row[0].row-10)).coordinate].value) & (portfolio['DIVISA']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() if portfolio.loc[(portfolio['INTERMEDIARIO']==ws[row[_].offset(column=0, row=-2-(row[0].row-10)).coordinate].value) & (portfolio['DIVISA']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() != 0 else ''
                    ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                    ws[row[_].coordinate].font = Font(name='Times New Roman', size=9)
                    ws[row[_].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                    ws[row[_].coordinate].number_format = '#,0'

                ws[row[num_intermediari+1].coordinate].value = portfolio.loc[portfolio['DIVISA']==ws[row[0].coordinate].value, 'TOTALE t1'].sum()
                ws[row[num_intermediari+1].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+1].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+1].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+1].coordinate].number_format = '#,0'
                ws[row[num_intermediari+2].coordinate].value = portfolio.loc[portfolio['DIVISA']==ws[row[0].coordinate].value, 'TOTALE t0'].sum()
                ws[row[num_intermediari+2].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+2].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+2].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+2].coordinate].number_format = '#,0'

                #ws[row[0].coordinate].value = tipo_divisa_dict[ws[row[0].coordinate].value] # aggiorna valori dell'indice con i nomi nel dizionario
                lunghezza_colonna_18.append(len(ws.cell(row=row[0].row, column=row[0].column).value)) # ottieni la lunghezza della colonna
                ws.column_dimensions[row[0].column_letter].width = max(lunghezza_colonna_18) + 7.5 # modifica larghezza colonna


        # Somma per intermediari
        for row in ws.iter_rows(min_row=10 + len_tipo_divisa, max_row=10 + len_tipo_divisa, min_col=min_col, max_col=min_col + len_header_18):
            ws[row[0].coordinate].value = 'TOTALE'
            ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
            ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
            ws[row[0].coordinate].border = Border(bottom=Side(border_style='thin', color='31869B'), right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'), top=Side(border_style='thin', color='31869B'))
            ws.row_dimensions[row[0].row].height = 19
            for _ in range(1,len_header_18-2):
                ws[row[_].coordinate].value = portfolio.loc[portfolio['INTERMEDIARIO']==ws.cell(row=row[_].row, column=row[_].column).offset(row=-len_tipo_divisa-2).value, 'TOTALE t1'].sum()
            ws[row[len_header_18-2].coordinate].value = portfolio.loc[:, 'TOTALE t1'].sum()
            ws[row[len_header_18-1].coordinate].value = portfolio.loc[:, 'TOTALE t0'].sum()
            for _ in range(1,len_header_18):
                ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                ws[row[_].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
                ws[row[_].coordinate].border = Border(bottom=Side(border_style='thin', color='31869B'), right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'), top=Side(border_style='thin', color='31869B'))
                ws[row[_].coordinate].number_format = '#,0'

    def tabella_pivot_azioni(self):
        """
        Crea la tabella pivot delle azioni.
        """
        # Carica portafoglio
        portfolio = pd.read_excel(self.file_portafoglio, sheet_name='Portfolio', header=1)

        ws = self.wb.create_sheet('19.azioni')
        ws = self.wb['19.azioni']
        self.wb.active = ws

        # Inizializza variabili
        min_row = 1
        ptf_equity = portfolio.loc[portfolio['CATEGORIA']=='EQUITY']
        banks = ptf_equity['INTERMEDIARIO'].unique()
        count_banks = banks.size
        dict_equity = ptf_equity.to_dict('list')
        
        # Creazione tabella #
        
        # Colonne
        header = banks.tolist()
        header.insert(0, '')
        header.extend(('Totale '+ self.mesi_dict[self.t1.month], 'Totale '+ self.mesi_dict[self.t0_1m.month], 'Delta'))

        # Titolo
        ws['A1'] = 'Azioni'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        match count_banks:
            case 1:
                lunghezza_titolo = 12
                min_col = 4
            case 2:
                lunghezza_titolo = 12
                min_col = 4
            case 3:
                lunghezza_titolo = 12
                min_col = 3
            case 4:
                lunghezza_titolo = 12
                min_col = 3
            case 5:
                lunghezza_titolo = 12
                min_col = 2
            case 6:
                lunghezza_titolo = 12
                min_col = 2
            case 7:
                lunghezza_titolo = 12
                min_col = 1
            case _:
                lunghezza_titolo = len(header)
                min_col = 1
        ws.merge_cells(start_row=1, end_row=4, start_column=1, end_column=lunghezza_titolo)
        min_row += 7

        # Intestazione
        for col in ws.iter_cols(min_row=min_row, max_row=min_row + 1, min_col=min_col, max_col=min_col + len(header) - 1):
            ws[col[0].coordinate].value = header[col[0].column-min_col]
            ws[col[0].coordinate].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col[0].coordinate].font = Font(name='Times New Roman', size=10, color='FFFFFF')
            ws[col[0].coordinate].fill = PatternFill(fill_type='solid', fgColor='31869B')
            ws[col[0].coordinate].border = Border(
                right=Side(border_style='thin', color='31869B'),
                left=Side(border_style='thin', color='31869B')
            )
            ws.merge_cells(start_row=col[0].row, end_row=col[1].row, start_column=col[0].column, end_column=col[0].column)
            ws.row_dimensions[col[0].row].height = 20
            ws.row_dimensions[col[1].row].height = 20
            ws.column_dimensions[col[0].column_letter].width = 12
        min_row += 2

        # Corpo
        count_equity = len(dict_equity['PRODOTTO'])
        columns_length = []
        for row in ws.iter_rows(min_row=min_row, max_row=min_row + count_equity - 1, min_col=min_col, max_col=min_col + len(header)):
            # nome strumento
            ws[row[0].coordinate].value = dict_equity['PRODOTTO'][row[0].row-min_row]
            ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
            ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, color='000000')
            ws[row[0].coordinate].border = Border(
                bottom=Side(border_style='dashed', color='31869B'),
                right=Side(border_style='dashed', color='31869B'),
                left=Side(border_style='dashed', color='31869B')
            )
            # controvalore strumento aggiunto per riga
            for _ in range(1, count_banks+1):
                row_offset = -2-(row[0].row-min_row)
                ctv_t1 = ptf_equity.loc[
                    (ptf_equity['INTERMEDIARIO']==ws[row[_].offset(column=0, row=row_offset).coordinate].value) &
                    (ptf_equity['PRODOTTO']==ws[row[0].coordinate].value),
                    'TOTALE t1'
                ]
                if not ctv_t1.empty:
                    ws[row[_].coordinate].value = ctv_t1.values[0]
                else:
                    ws[row[_].coordinate].value = ''
                ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                ws[row[_].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[_].coordinate].border = Border(
                    bottom=Side(border_style='dashed', color='31869B'),
                    right=Side(border_style='dashed', color='31869B'),
                    left=Side(border_style='dashed', color='31869B')
                )
                ws[row[_].coordinate].number_format = '#,0'
            # totale t1
            ws[row[count_banks+1].coordinate].value = dict_equity['TOTALE t1'][row[0].row-min_row]
            ws[row[count_banks+1].coordinate].alignment = Alignment(horizontal='center')
            ws[row[count_banks+1].coordinate].font = Font(name='Times New Roman', size=9)
            ws[row[count_banks+1].coordinate].border = Border(
                bottom=Side(border_style='dashed', color='31869B'),
                right=Side(border_style='dashed', color='31869B'),
                left=Side(border_style='dashed', color='31869B')
            )
            ws[row[count_banks+1].coordinate].number_format = '#,0'
            # totale t0
            ws[row[count_banks+2].coordinate].value = dict_equity['TOTALE t0'][row[0].row-min_row]
            ws[row[count_banks+2].coordinate].alignment = Alignment(horizontal='center')
            ws[row[count_banks+2].coordinate].font = Font(name='Times New Roman', size=9)
            ws[row[count_banks+2].coordinate].border = Border(
                bottom=Side(border_style='dashed', color='31869B'),
                right=Side(border_style='dashed', color='31869B'),
                left=Side(border_style='dashed', color='31869B')
            )
            ws[row[count_banks+2].coordinate].number_format = '#,0'
            # calcolo del delta mensile dei prezzi degli strumenti in euro,
            # tranne quando il prodotto è stato liquidato in t1 o non esisteva in t0
            if ws[row[count_banks+1].coordinate].value != 0 and ws[row[count_banks+2].coordinate].value != 0:
                ws[row[count_banks+3].coordinate].value = (
                    (
                        (dict_equity['PREZZO t1'][row[0].row-min_row] * dict_equity['CAMBIO t1'][row[0].row-min_row])
                        - (dict_equity['PREZZO t0'][row[0].row-min_row]* dict_equity['CAMBIO t0'][row[0].row-min_row])
                    )
                    /
                    (dict_equity['PREZZO t0'][row[0].row-min_row] * dict_equity['CAMBIO t0'][row[0].row-min_row])
                )
            else:
                ws[row[count_banks+3].coordinate].value = '/'
            ws[row[count_banks+3].coordinate].alignment = Alignment(horizontal='center')
            ws[row[count_banks+3].coordinate].font = Font(name='Times New Roman', size=9)
            ws[row[count_banks+3].coordinate].border = Border(
                bottom=Side(border_style='dashed', color='31869B'),
                right=Side(border_style='dashed', color='31869B'),
                left=Side(border_style='dashed', color='31869B')
            )
            ws[row[count_banks+3].coordinate].number_format = FORMAT_PERCENTAGE_00

            columns_length.append(len(ws.cell(row=row[0].row, column=row[0].column).value)) # ottieni la lunghezza della colonna
            ws.column_dimensions[row[0].column_letter].width = max(columns_length) + 2.5 # modifica larghezza colonna

        # Somma per intermediari
        for row in ws.iter_rows(min_row=min_row + count_equity, max_row=min_row + count_equity, min_col=min_col, max_col=min_col + len(header)):
            ws[row[0].coordinate].value = 'TOTALE'
            ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
            ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
            ws[row[0].coordinate].border = Border(
                bottom=Side(border_style='thin', color='31869B'),
                right=Side(border_style='thin', color='31869B'),
                left=Side(border_style='thin', color='31869B'),
                top=Side(border_style='thin', color='31869B')
            )
            for _ in range(1, len(header)-3):
                ws[row[_].coordinate].value = ptf_equity.loc[
                    ptf_equity['INTERMEDIARIO']==ws.cell(row=row[_].row, column=row[_].column).offset(row=-count_equity-2).value,
                    'TOTALE t1'
                ].sum()
            ws[row[len(header)-3].coordinate].value = ptf_equity['TOTALE t1'].sum()
            ws[row[len(header)-2].coordinate].value = ptf_equity['TOTALE t0'].sum()
            # delta mensile complessivo
            ptf_equity_not_null = ptf_equity.loc[(ptf_equity['TOTALE t1']!=0) & (ptf_equity['TOTALE t0']!=0)]

            ws[row[len(header)-1].coordinate].value = (
                # somma dei controvalori t1 ottenuti con quantità vecchie
                (
                    # prezzi nuovi in euro
                    (ptf_equity_not_null['PREZZO t1'] * ptf_equity_not_null['CAMBIO t1'])
                    # moltiplicati per le quantità vecchie
                    *
                    ptf_equity_not_null['QUANTITA t0']
                ).sum()
                # divisi per la somma dei controvalori t0 con quantità vecchie
                /
                (
                    # prezzi vecchi in euro
                    (ptf_equity_not_null['PREZZO t0'] * ptf_equity_not_null['CAMBIO t0'])
                    # moltiplicati per le quantità vecchie
                    *
                    ptf_equity_not_null['QUANTITA t0']
                ).sum()
            ) - 1

            for _ in range(1, len(header)):
                ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                ws[row[_].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
                ws[row[_].coordinate].border = Border(
                    bottom=Side(border_style='thin', color='31869B'),
                    right=Side(border_style='thin', color='31869B'),
                    left=Side(border_style='thin', color='31869B'),
                    top=Side(border_style='thin', color='31869B'))
                ws[row[_].coordinate].number_format = '#,0'
            ws[row[len(header)-1].coordinate].number_format = FORMAT_PERCENTAGE_00

    def tabella_pivot_obbligazioni_governative(self):
        """
        Crea la tabella pivot delle obbligazioni governative.
        """
        # Carica portafoglio
        portfolio = pd.read_excel(self.file_portafoglio, sheet_name='Portfolio', header=1)

        ws = self.wb.create_sheet('20.obb_gov')
        ws = self.wb['20.obb_gov']
        self.wb.active = ws

        # Inizializza variabili
        min_row = 1
        ptf_gov_bond = portfolio.loc[portfolio['CATEGORIA']=='GOVERNMENT_BOND']
        banks = ptf_gov_bond['INTERMEDIARIO'].unique()
        count_banks = banks.size
        dict_gov_bond = ptf_gov_bond.to_dict('list')

        # Creazione tabella #
        
        # Colonne
        header = banks.tolist()
        header.insert(0, '')
        header.extend(('Totale '+ self.mesi_dict[self.t1.month], 'Totale '+ self.mesi_dict[self.t0_1m.month], 'Delta'))

        # Creazione tabella # TODO da togliere
        header_20 = list(portfolio.loc[portfolio['CATEGORIA']=='GOVERNMENT_BOND','INTERMEDIARIO'].unique())
        header_20.insert(0, '')
        header_20.extend(('Totale '+ self.mesi_dict[self.t1.month], 'Totale '+ self.mesi_dict[self.t0_1m.month], 'Delta'))
        len_header_20 = len(header_20)

        ws['A1'] = 'Obbligazioni Governative'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        match count_banks:
            case 1:
                lunghezza_titolo = 12
                min_col = 4
            case 2:
                lunghezza_titolo = 12
                min_col = 4
            case 3:
                lunghezza_titolo = 12
                min_col = 3
            case 4:
                lunghezza_titolo = 12
                min_col = 3
            case 5:
                lunghezza_titolo = 12
                min_col = 2
            case 6:
                lunghezza_titolo = 12
                min_col = 2
            case 7:
                lunghezza_titolo = 12
                min_col = 1
            case _:
                lunghezza_titolo = len(header)
                min_col = 1
        ws.merge_cells(start_row=1, end_row=4, start_column=1, end_column=lunghezza_titolo)
        min_row += 7

        # Intestazione
        for col in ws.iter_cols(min_row=min_row, max_row=min_row + 1, min_col=min_col, max_col=min_col + len(header) - 1):
            ws[col[0].coordinate].value = header[col[0].column-min_col]
            ws[col[0].coordinate].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col[0].coordinate].font = Font(name='Times New Roman', size=10, color='FFFFFF')
            ws[col[0].coordinate].fill = PatternFill(fill_type='solid', fgColor='31869B')
            ws[col[0].coordinate].border = Border(
                right=Side(border_style='thin', color='31869B'),
                left=Side(border_style='thin', color='31869B')
            )
            ws.merge_cells(start_row=col[0].row, end_row=col[1].row, start_column=col[0].column, end_column=col[0].column)
            ws.row_dimensions[col[0].row].height = 20
            ws.row_dimensions[col[1].row].height = 20
            ws.column_dimensions[col[0].column_letter].width = 12
        min_row += 2

        # Corpo
        count_gov_bond = len(dict_gov_bond['PRODOTTO'])
        columns_length = []
        nome_obbgov = list(portfolio.loc[portfolio['CATEGORIA']=='GOVERNMENT_BOND','PRODOTTO'])
        len_nome_obbgov = len(nome_obbgov)
        num_intermediari = len(portfolio.loc[portfolio['CATEGORIA']=='GOVERNMENT_BOND', 'INTERMEDIARIO'].unique())
        for row in ws.iter_rows(min_row=min_row, max_row=min_row + count_gov_bond - 1, min_col=min_col, max_col=min_col + len(header)):
            # nome strumento
            ws[row[0].coordinate].value = dict_gov_bond['PRODOTTO'][row[0].row-min_row]
            ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
            ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, color='000000')
            ws[row[0].coordinate].border = Border(
                bottom=Side(border_style='dashed', color='31869B'),
                right=Side(border_style='dashed', color='31869B'),
                left=Side(border_style='dashed', color='31869B')
            )
            # controvalore strumento aggiunto per riga
            for _ in range(1, count_banks+1):
                row_offset = -2-(row[0].row-min_row)
                ctv_t1 = ptf_gov_bond.loc[
                    (ptf_gov_bond['INTERMEDIARIO']==ws[row[_].offset(column=0, row=row_offset).coordinate].value) &
                    (ptf_gov_bond['PRODOTTO']==ws[row[0].coordinate].value),
                    'TOTALE t1'
                ]
                if not ctv_t1.empty:
                    ws[row[_].coordinate].value = ctv_t1.values[0]
                else:
                    ws[row[_].coordinate].value = ''
                ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                ws[row[_].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[_].coordinate].border = Border(
                    bottom=Side(border_style='dashed', color='31869B'),
                    right=Side(border_style='dashed', color='31869B'),
                    left=Side(border_style='dashed', color='31869B')
                )
                ws[row[_].coordinate].number_format = '#,0'
            # totale t1
            ws[row[count_banks+1].coordinate].value = dict_gov_bond['TOTALE t1'][row[0].row-min_row]
            ws[row[count_banks+1].coordinate].alignment = Alignment(horizontal='center')
            ws[row[count_banks+1].coordinate].font = Font(name='Times New Roman', size=9)
            ws[row[count_banks+1].coordinate].border = Border(
                bottom=Side(border_style='dashed', color='31869B'),
                right=Side(border_style='dashed', color='31869B'),
                left=Side(border_style='dashed', color='31869B')
            )
            ws[row[count_banks+1].coordinate].number_format = '#,0'
            # totale t0
            ws[row[count_banks+2].coordinate].value = dict_gov_bond['TOTALE t0'][row[0].row-min_row]
            ws[row[count_banks+2].coordinate].alignment = Alignment(horizontal='center')
            ws[row[count_banks+2].coordinate].font = Font(name='Times New Roman', size=9)
            ws[row[count_banks+2].coordinate].border = Border(
                bottom=Side(border_style='dashed', color='31869B'),
                right=Side(border_style='dashed', color='31869B'),
                left=Side(border_style='dashed', color='31869B')
            )
            ws[row[count_banks+2].coordinate].number_format = '#,0'
            # calcolo del delta mensile dei prezzi degli strumenti in euro,
            # tranne quando il prodotto è stato liquidato in t1 o non esisteva in t0
            if ws[row[count_banks+1].coordinate].value != 0 and ws[row[count_banks+2].coordinate].value != 0:
                ws[row[count_banks+3].coordinate].value = (
                    (
                        (dict_gov_bond['PREZZO t1'][row[0].row-min_row] * dict_gov_bond['CAMBIO t1'][row[0].row-min_row])
                        - (dict_gov_bond['PREZZO t0'][row[0].row-min_row]* dict_gov_bond['CAMBIO t0'][row[0].row-min_row])
                    )
                    /
                    (dict_gov_bond['PREZZO t0'][row[0].row-min_row] * dict_gov_bond['CAMBIO t0'][row[0].row-min_row])
                )
            else:
                ws[row[count_banks+3].coordinate].value = '/'
            ws[row[count_banks+3].coordinate].alignment = Alignment(horizontal='center')
            ws[row[count_banks+3].coordinate].font = Font(name='Times New Roman', size=9)
            ws[row[count_banks+3].coordinate].border = Border(
                bottom=Side(border_style='dashed', color='31869B'),
                right=Side(border_style='dashed', color='31869B'),
                left=Side(border_style='dashed', color='31869B')
            )
            ws[row[count_banks+3].coordinate].number_format = FORMAT_PERCENTAGE_00

            columns_length.append(len(ws.cell(row=row[0].row, column=row[0].column).value)) # ottieni la lunghezza della colonna
            ws.column_dimensions[row[0].column_letter].width = max(columns_length) + 2.5 # modifica larghezza colonna FALLO ALLA FINE

        # Somma per intermediari
        for row in ws.iter_rows(min_row=min_row + count_gov_bond, max_row=min_row + count_gov_bond, min_col=min_col, max_col=min_col + len(header)):
            ws[row[0].coordinate].value = 'TOTALE'
            ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
            ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
            ws[row[0].coordinate].border = Border(
                bottom=Side(border_style='thin', color='31869B'),
                right=Side(border_style='thin', color='31869B'),
                left=Side(border_style='thin', color='31869B'),
                top=Side(border_style='thin', color='31869B')
            )
            for _ in range(1, len(header)-3):
                # TODO: da qui!
                ws[row[_].coordinate].value = portfolio.loc[(portfolio['INTERMEDIARIO']==ws.cell(row=row[_].row, column=row[_].column).offset(row=-len_nome_obbgov-2).value) & (portfolio['CATEGORIA']=='GOVERNMENT_BOND'), 'TOTALE t1'].sum()
            ws[row[len_header_20-3].coordinate].value = portfolio.loc[portfolio['CATEGORIA']=='GOVERNMENT_BOND', 'TOTALE t1'].sum()
            ws[row[len_header_20-2].coordinate].value = portfolio.loc[portfolio['CATEGORIA']=='GOVERNMENT_BOND', 'TOTALE t0'].sum()
            ws[row[len_header_20-1].coordinate].value = (portfolio.loc[(portfolio['CATEGORIA']=='GOVERNMENT_BOND') & (portfolio['TOTALE t0']!=0), 'TOTALE t1'].sum() - portfolio.loc[(portfolio['CATEGORIA']=='GOVERNMENT_BOND') & (portfolio['TOTALE t1']!=0), 'TOTALE t0'].sum()) / portfolio.loc[(portfolio['CATEGORIA']=='GOVERNMENT_BOND') & (portfolio['TOTALE t1']!=0), 'TOTALE t0'].sum()
            
            for _ in range(1, len(header)):
                ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                ws[row[_].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
                ws[row[_].coordinate].border = Border(
                    bottom=Side(border_style='thin', color='31869B'),
                    right=Side(border_style='thin', color='31869B'),
                    left=Side(border_style='thin', color='31869B'),
                    top=Side(border_style='thin', color='31869B')
                )
                ws[row[_].coordinate].number_format = '#,0'
            ws[row[len(header)-1].coordinate].number_format = FORMAT_PERCENTAGE_00

    def tabella_pivot_obbligazioni_societarie(self):
        """
        Crea la tabella pivot delle obbligazioni societarie.
        """
        # Carica portafoglio
        portfolio = pd.read_excel(self.file_portafoglio, sheet_name='Portfolio', header=1)

        ws = self.wb.create_sheet('21.obb_cor')
        ws = self.wb['21.obb_cor']
        self.wb.active = ws

        # Inizializza variabili
        min_row = 1
        ptf_corp_bond = portfolio.loc[portfolio['CATEGORIA']=='CORPORATE_BOND']
        banks = ptf_corp_bond['INTERMEDIARIO'].unique()
        count_banks = banks.size
        dict_corp_bond = ptf_corp_bond.to_dict('list')

        # Creazione tabella #
        
        # Colonne
        header = banks.tolist()
        header.insert(0, '')
        header.extend(('Totale '+ self.mesi_dict[self.t1.month], 'Totale '+ self.mesi_dict[self.t0_1m.month], 'Delta'))

        # Creazione tabella # TODO da togliere
        header_21 = list(portfolio.loc[portfolio['CATEGORIA']=='CORPORATE_BOND','INTERMEDIARIO'].unique())
        header_21.insert(0, '')
        header_21.extend(('Totale '+ self.mesi_dict[self.t1.month], 'Totale '+ self.mesi_dict[self.t0_1m.month], 'Delta'))
        len_header_21 = len(header_21)

        ws['A1'] = 'Obbligazioni Corporate'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        match count_banks:
            case 1:
                lunghezza_titolo = 12
                min_col = 4
            case 2:
                lunghezza_titolo = 12
                min_col = 4
            case 3:
                lunghezza_titolo = 12
                min_col = 3
            case 4:
                lunghezza_titolo = 12
                min_col = 3
            case 5:
                lunghezza_titolo = 12
                min_col = 2
            case 6:
                lunghezza_titolo = 12
                min_col = 2
            case 7:
                lunghezza_titolo = 12
                min_col = 1
            case _:
                lunghezza_titolo = len(header)
                min_col = 1
        ws.merge_cells(start_row=1, end_row=4, start_column=1, end_column=lunghezza_titolo)
        min_row += 7

        # Intestazione
        for col in ws.iter_cols(min_row=min_row, max_row=min_row + 1, min_col=min_col, max_col=min_col + len(header) - 1):
            ws[col[0].coordinate].value = header[col[0].column-min_col]
            ws[col[0].coordinate].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col[0].coordinate].font = Font(name='Times New Roman', size=10, color='FFFFFF')
            ws[col[0].coordinate].fill = PatternFill(fill_type='solid', fgColor='31869B')
            ws[col[0].coordinate].border = Border(
                right=Side(border_style='thin', color='31869B'),
                left=Side(border_style='thin', color='31869B')
            )
            ws.merge_cells(start_row=col[0].row, end_row=col[1].row, start_column=col[0].column, end_column=col[0].column)
            ws.row_dimensions[col[0].row].height = 20
            ws.row_dimensions[col[1].row].height = 20
            ws.column_dimensions[col[0].column_letter].width = 12
        min_row += 2

        # Corpo
        count_corp_bond = len(dict_corp_bond['PRODOTTO'])
        columns_length = []
        nome_obbcorp = list(portfolio.loc[portfolio['CATEGORIA']=='CORPORATE_BOND','PRODOTTO'])
        len_nome_obbcorp = len(nome_obbcorp)
        num_intermediari = len(portfolio.loc[portfolio['CATEGORIA']=='CORPORATE_BOND', 'INTERMEDIARIO'].unique())
        for row in ws.iter_rows(min_row=min_row, max_row=min_row + count_corp_bond - 1, min_col=min_col, max_col=min_col + len(header)):
            # nome strumento
            ws[row[0].coordinate].value = dict_corp_bond['PRODOTTO'][row[0].row-min_row]
            ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
            ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, color='000000')
            ws[row[0].coordinate].border = Border(
                bottom=Side(border_style='dashed', color='31869B'),
                right=Side(border_style='dashed', color='31869B'),
                left=Side(border_style='dashed', color='31869B')
            )
            # controvalore strumento aggiunto per riga
            for _ in range(1, count_banks+1):
                row_offset = -2-(row[0].row-min_row)
                ctv_t1 = ptf_corp_bond.loc[
                    (ptf_corp_bond['INTERMEDIARIO']==ws[row[_].offset(column=0, row=row_offset).coordinate].value) &
                    (ptf_corp_bond['PRODOTTO']==ws[row[0].coordinate].value),
                    'TOTALE t1'
                ]
                if not ctv_t1.empty:
                    ws[row[_].coordinate].value = ctv_t1.values[0]
                else:
                    ws[row[_].coordinate].value = ''
                ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                ws[row[_].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[_].coordinate].border = Border(
                    bottom=Side(border_style='dashed', color='31869B'),
                    right=Side(border_style='dashed', color='31869B'),
                    left=Side(border_style='dashed', color='31869B')
                )
                ws[row[_].coordinate].number_format = '#,0'
            # totale t1
            ws[row[count_banks+1].coordinate].value = dict_corp_bond['TOTALE t1'][row[0].row-min_row]
            ws[row[count_banks+1].coordinate].alignment = Alignment(horizontal='center')
            ws[row[count_banks+1].coordinate].font = Font(name='Times New Roman', size=9)
            ws[row[count_banks+1].coordinate].border = Border(
                bottom=Side(border_style='dashed', color='31869B'),
                right=Side(border_style='dashed', color='31869B'),
                left=Side(border_style='dashed', color='31869B')
            )
            ws[row[count_banks+1].coordinate].number_format = '#,0'
            # totale t0
            ws[row[count_banks+2].coordinate].value = dict_corp_bond['TOTALE t0'][row[0].row-min_row]
            ws[row[count_banks+2].coordinate].alignment = Alignment(horizontal='center')
            ws[row[count_banks+2].coordinate].font = Font(name='Times New Roman', size=9)
            ws[row[count_banks+2].coordinate].border = Border(
                bottom=Side(border_style='dashed', color='31869B'),
                right=Side(border_style='dashed', color='31869B'),
                left=Side(border_style='dashed', color='31869B')
            )
            ws[row[count_banks+2].coordinate].number_format = '#,0'
            # calcolo del delta mensile dei prezzi degli strumenti in euro,
            # tranne quando il prodotto è stato liquidato in t1 o non esisteva in t0
            if ws[row[count_banks+1].coordinate].value != 0 and ws[row[count_banks+2].coordinate].value != 0:
                ws[row[count_banks+3].coordinate].value = (
                    (
                        (dict_corp_bond['PREZZO t1'][row[0].row-min_row] * dict_corp_bond['CAMBIO t1'][row[0].row-min_row])
                        - (dict_corp_bond['PREZZO t0'][row[0].row-min_row]* dict_corp_bond['CAMBIO t0'][row[0].row-min_row])
                    )
                    /
                    (dict_corp_bond['PREZZO t0'][row[0].row-min_row] * dict_corp_bond['CAMBIO t0'][row[0].row-min_row])
                )
            else:
                ws[row[count_banks+3].coordinate].value = '/'
            ws[row[count_banks+3].coordinate].alignment = Alignment(horizontal='center')
            ws[row[count_banks+3].coordinate].font = Font(name='Times New Roman', size=9)
            ws[row[count_banks+3].coordinate].border = Border(
                bottom=Side(border_style='dashed', color='31869B'),
                right=Side(border_style='dashed', color='31869B'),
                left=Side(border_style='dashed', color='31869B')
            )
            ws[row[count_banks+3].coordinate].number_format = FORMAT_PERCENTAGE_00

            columns_length.append(len(ws.cell(row=row[0].row, column=row[0].column).value)) # ottieni la lunghezza della colonna
            ws.column_dimensions[row[0].column_letter].width = max(columns_length) + 2.5 # modifica larghezza colonna FALLO ALLA FINE

        # Somma per intermediari
        for row in ws.iter_rows(min_row=min_row + count_corp_bond, max_row=min_row + count_corp_bond, min_col=min_col, max_col=min_col + len(header)):
            ws[row[0].coordinate].value = 'TOTALE'
            ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
            ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
            ws[row[0].coordinate].border = Border(
                bottom=Side(border_style='thin', color='31869B'),
                right=Side(border_style='thin', color='31869B'),
                left=Side(border_style='thin', color='31869B'),
                top=Side(border_style='thin', color='31869B')
            )
            for _ in range(1, len(header)-3):
                # TODO: da qui!
                ws[row[_].coordinate].value = portfolio.loc[(portfolio['INTERMEDIARIO']==ws.cell(row=row[_].row, column=row[_].column).offset(row=-len_nome_obbcorp-2).value) & (portfolio['CATEGORIA']=='CORPORATE_BOND'), 'TOTALE t1'].sum()
            ws[row[len_header_21-3].coordinate].value = portfolio.loc[portfolio['CATEGORIA']=='CORPORATE_BOND', 'TOTALE t1'].sum()
            ws[row[len_header_21-2].coordinate].value = portfolio.loc[portfolio['CATEGORIA']=='CORPORATE_BOND', 'TOTALE t0'].sum()
            ws[row[len_header_21-1].coordinate].value = (portfolio.loc[(portfolio['CATEGORIA']=='CORPORATE_BOND') & (portfolio['TOTALE t0']!=0), 'TOTALE t1'].sum() - portfolio.loc[(portfolio['CATEGORIA']=='CORPORATE_BOND') & (portfolio['TOTALE t1']!=0), 'TOTALE t0'].sum()) / portfolio.loc[(portfolio['CATEGORIA']=='CORPORATE_BOND') & (portfolio['TOTALE t1']!=0), 'TOTALE t0'].sum()
            
            for _ in range(1, len(header)):
                ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                ws[row[_].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
                ws[row[_].coordinate].border = Border(
                    bottom=Side(border_style='thin', color='31869B'),
                    right=Side(border_style='thin', color='31869B'),
                    left=Side(border_style='thin', color='31869B'),
                    top=Side(border_style='thin', color='31869B')
                )
                ws[row[_].coordinate].number_format = '#,0'
            ws[row[len(header)-1].coordinate].number_format = FORMAT_PERCENTAGE_00

    def obb_totale_22(self):
        """
        Crea la ventiduesima pagina.
        """
        # Carica portafoglio
        portfolio = pd.read_excel(self.file_portafoglio, sheet_name='Portfolio', header=1)

        # 22.Obb. totale
        ws = self.wb.create_sheet('22.obb_tot')
        ws = self.wb['22.obb_tot']
        self.wb.active = ws

        # Creazione tabella
        header_22 = ['', 'OBBLIGAZIONE GOVERNATIVE', 'OBBLIGAZIONE CORPORATE', 'Totale '+ self.mesi_dict[self.t1.month], 'Totale '+ self.mesi_dict[self.t0_1m.month]]
        len_header_22 = len(header_22)

        # Titolo
        ws['A1'] = 'Riepilogo Obbligazioni'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells(start_row=1, end_row=4, start_column=1, end_column=12)
        min_col = 4

        # Intestazione
        for col in ws.iter_cols(min_row=8, max_row=9, min_col=min_col, max_col=min_col + len_header_22 -1):
            ws[col[0].coordinate].value = header_22[0]
            del header_22[0]
            ws[col[0].coordinate].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col[0].coordinate].font = Font(name='Times New Roman', size=10, color='FFFFFF')
            ws[col[0].coordinate].fill = PatternFill(fill_type='solid', fgColor='31869B')
            ws[col[0].coordinate].border = Border(right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'))
            ws.merge_cells(start_row=col[0].row, end_row=col[1].row, start_column=col[0].column, end_column=col[0].column)
            ws.row_dimensions[col[0].row].height = 20
            ws.row_dimensions[col[1].row].height = 20
            ws.column_dimensions[col[0].column_letter].width = 15

        # Indice e riempimento tabella
        int_obb = list(portfolio.loc[(portfolio['CATEGORIA']=='GOVERNMENT_BOND') | (portfolio['CATEGORIA']=='CORPORATE_BOND'),'INTERMEDIARIO'].unique())
        len_int_obb = len(int_obb)
        #num_intermediari = len(portfolio.loc[portfolio['CATEGORIA']=='CORPORATE_BOND', 'INTERMEDIARIO'].unique())
        num_intermediari = 2
        lunghezza_colonna_22 = []
        for row in ws.iter_rows(min_row=8, max_row=10 + len_int_obb -1, min_col=min_col, max_col=min_col + len_header_22):
            if row[0].row > 9:
                ws[row[0].coordinate].value = int_obb[0]
                del int_obb[0]
                ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
                ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, color='000000')
                ws[row[0].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))

                ws[row[1].coordinate].value = portfolio.loc[(portfolio['CATEGORIA']=='GOVERNMENT_BOND') & (portfolio['INTERMEDIARIO']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() if portfolio.loc[(portfolio['CATEGORIA']=='GOVERNMENT_BOND') & (portfolio['INTERMEDIARIO']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() != 0 else ''
                ws[row[1].coordinate].alignment = Alignment(horizontal='center')
                ws[row[1].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[1].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[1].coordinate].number_format = '#,0'
                ws[row[2].coordinate].value = portfolio.loc[(portfolio['CATEGORIA']=='CORPORATE_BOND') & (portfolio['INTERMEDIARIO']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() if portfolio.loc[(portfolio['CATEGORIA']=='CORPORATE_BOND') & (portfolio['INTERMEDIARIO']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() != 0 else ''
                ws[row[2].coordinate].alignment = Alignment(horizontal='center')
                ws[row[2].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[2].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[2].coordinate].number_format = '#,0'

                ws[row[num_intermediari+1].coordinate].value = portfolio.loc[(portfolio['INTERMEDIARIO']==ws[row[0].coordinate].value) & ((portfolio['CATEGORIA']=='GOVERNMENT_BOND') | (portfolio['CATEGORIA']=='CORPORATE_BOND')), 'TOTALE t1'].sum()
                ws[row[num_intermediari+1].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+1].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+1].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+1].coordinate].number_format = '#,0'
                ws[row[num_intermediari+2].coordinate].value = portfolio.loc[(portfolio['INTERMEDIARIO']==ws[row[0].coordinate].value) & ((portfolio['CATEGORIA']=='GOVERNMENT_BOND') | (portfolio['CATEGORIA']=='CORPORATE_BOND')), 'TOTALE t0'].sum()
                ws[row[num_intermediari+2].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+2].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+2].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+2].coordinate].number_format = '#,0'

                lunghezza_colonna_22.append(len(ws.cell(row=row[0].row, column=row[0].column).value)) # ottieni la lunghezza della colonna
                ws.column_dimensions[row[0].column_letter].width = max(lunghezza_colonna_22) + 2.5 # modifica larghezza colonna FALLO ALLA FINE

        # Somma per strumento
        for row in ws.iter_rows(min_row=10 + len_int_obb, max_row=10 + len_int_obb, min_col=min_col, max_col=min_col + len_header_22):
            ws[row[0].coordinate].value = 'TOTALE'
            ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
            ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
            ws[row[0].coordinate].border = Border(bottom=Side(border_style='thin', color='31869B'), right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'), top=Side(border_style='thin', color='31869B'))
            
            ws[row[1].coordinate].value = portfolio.loc[portfolio['CATEGORIA']=='GOVERNMENT_BOND', 'TOTALE t1'].sum()
            ws[row[2].coordinate].value = portfolio.loc[portfolio['CATEGORIA']=='CORPORATE_BOND', 'TOTALE t1'].sum()
            
            ws[row[len_header_22-2].coordinate].value = portfolio.loc[(portfolio['CATEGORIA']=='GOVERNMENT_BOND') | (portfolio['CATEGORIA']=='CORPORATE_BOND'), 'TOTALE t1'].sum()
            ws[row[len_header_22-1].coordinate].value = portfolio.loc[(portfolio['CATEGORIA']=='GOVERNMENT_BOND') | (portfolio['CATEGORIA']=='CORPORATE_BOND'), 'TOTALE t0'].sum()

            for _ in range(1,len_header_22):
                ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                ws[row[_].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
                ws[row[_].coordinate].border = Border(bottom=Side(border_style='thin', color='31869B'), right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'), top=Side(border_style='thin', color='31869B'))
                ws[row[_].coordinate].number_format = '#,0'

    def liquidità_23(self):
        """
        Crea la ventitreesima pagina.
        """
        # Carica portafoglio
        portfolio = pd.read_excel(self.file_portafoglio, sheet_name='Portfolio', header=1)

        ws = self.wb.create_sheet('23.liq')
        ws = self.wb['23.liq']
        self.wb.active = ws

        # Creazione tabella
        header_23 = list(portfolio.loc[(portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR'),'INTERMEDIARIO'].unique())
        header_23.insert(0, '')
        header_23.extend(('Totale '+ self.mesi_dict[self.t1.month], 'Totale '+ self.mesi_dict[self.t0_1m.month]))
        len_header_23 = len(header_23)

        # Titolo
        ws['A1'] = 'Liquidità'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        if len(list(portfolio.loc[(portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR'),'INTERMEDIARIO'].unique())) == 1:
            lunghezza_titolo_23 = 12
            min_col = 4
        elif len(list(portfolio.loc[(portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR'),'INTERMEDIARIO'].unique())) == 2:
            lunghezza_titolo_23 = 12
            min_col = 4
        elif len(list(portfolio.loc[(portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR'),'INTERMEDIARIO'].unique())) == 3:
            lunghezza_titolo_23 = 12
            min_col = 3
        elif len(list(portfolio.loc[(portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR'),'INTERMEDIARIO'].unique())) == 4:
            lunghezza_titolo_23 = 12
            min_col = 3
        elif len(list(portfolio.loc[(portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR'),'INTERMEDIARIO'].unique())) == 5:
            lunghezza_titolo_23 = 12
            min_col = 2
        elif len(list(portfolio.loc[(portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR'),'INTERMEDIARIO'].unique())) == 6:
            lunghezza_titolo_23 = 12
            min_col = 2
        elif len(list(portfolio.loc[(portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR'),'INTERMEDIARIO'].unique())) == 7:
            lunghezza_titolo_23 = 12
            min_col = 1
        else:
            lunghezza_titolo_23 = len_header_23
            min_col = 1
        ws.merge_cells(start_row=1, end_row=4, start_column=1, end_column=lunghezza_titolo_23)

        # Intestazione
        for col in ws.iter_cols(min_row=8, max_row=9, min_col=min_col, max_col=min_col + len_header_23 - 1):
            ws[col[0].coordinate].value = header_23[0]
            del header_23[0]
            ws[col[0].coordinate].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col[0].coordinate].font = Font(name='Times New Roman', size=10, color='FFFFFF')
            ws[col[0].coordinate].fill = PatternFill(fill_type='solid', fgColor='31869B')
            ws[col[0].coordinate].border = Border(right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'))
            ws.merge_cells(start_row=col[0].row, end_row=col[1].row, start_column=col[0].column, end_column=col[0].column)
            ws.row_dimensions[col[0].row].height = 20
            ws.row_dimensions[col[1].row].height = 20
            ws.column_dimensions[col[0].column_letter].width = 12

        # Indice e riempimento tabella
        nome_liq = list(portfolio.loc[(portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR'),'PRODOTTO'])
        len_nome_liq = len(nome_liq)
        num_intermediari = len(portfolio.loc[(portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR'), 'INTERMEDIARIO'].unique())
        lunghezza_colonna_23 = []
        for row in ws.iter_rows(min_row=8, max_row=10 + len_nome_liq -1, min_col=min_col, max_col=min_col + len_header_23):
            if row[0].row > 9:
                ws[row[0].coordinate].value = nome_liq[0]
                del nome_liq[0]
                ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
                ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, color='000000')
                ws[row[0].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                for _ in range(1, num_intermediari+1):
                    ws[row[_].coordinate].value = portfolio.loc[(portfolio['INTERMEDIARIO']==ws[row[_].offset(column=0, row=-2-(row[0].row-10)).coordinate].value) & (portfolio['PRODOTTO']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() if portfolio.loc[(portfolio['INTERMEDIARIO']==ws[row[_].offset(column=0, row=-2-(row[0].row-10)).coordinate].value) & (portfolio['PRODOTTO']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() != 0 else ''
                    ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                    ws[row[_].coordinate].font = Font(name='Times New Roman', size=9)
                    ws[row[_].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                    ws[row[_].coordinate].number_format = '#,0'

                ws[row[num_intermediari+1].coordinate].value = portfolio.loc[portfolio['PRODOTTO']==ws[row[0].coordinate].value, 'TOTALE t1'].sum()
                ws[row[num_intermediari+1].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+1].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+1].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+1].coordinate].number_format = '#,0'
                ws[row[num_intermediari+2].coordinate].value = portfolio.loc[portfolio['PRODOTTO']==ws[row[0].coordinate].value, 'TOTALE t0'].sum()
                ws[row[num_intermediari+2].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+2].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+2].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+2].coordinate].number_format = '#,0'

                lunghezza_colonna_23.append(len(ws.cell(row=row[0].row, column=row[0].column).value)) # ottieni la lunghezza della colonna
                ws.column_dimensions[row[0].column_letter].width = max(lunghezza_colonna_23) + 2.5 # modifica larghezza colonna

        # Somma per intermediari
        for row in ws.iter_rows(min_row=10 + len_nome_liq, max_row=10 + len_nome_liq, min_col=min_col, max_col=min_col + len_header_23):
            ws[row[0].coordinate].value = 'TOTALE'
            ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
            ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
            ws[row[0].coordinate].border = Border(bottom=Side(border_style='thin', color='31869B'), right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'), top=Side(border_style='thin', color='31869B'))
            for _ in range(1,len_header_23-2):
                ws[row[_].coordinate].value = portfolio.loc[(portfolio['INTERMEDIARIO']==ws.cell(row=row[_].row, column=row[_].column).offset(row=-len_nome_liq-2).value) & ((portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR')), 'TOTALE t1'].sum()
            ws[row[len_header_23-2].coordinate].value = portfolio.loc[(portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR'), 'TOTALE t1'].sum()
            ws[row[len_header_23-1].coordinate].value = portfolio.loc[(portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR'), 'TOTALE t0'].sum()
            for _ in range(1,len_header_23):
                ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                ws[row[_].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
                ws[row[_].coordinate].border = Border(bottom=Side(border_style='thin', color='31869B'), right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'), top=Side(border_style='thin', color='31869B'))
                ws[row[_].coordinate].number_format = '#,0'

    def liq_totale_24(self):
        """
        Crea la ventiquattresima pagina.
        """
        # Carica portafoglio
        portfolio = pd.read_excel(self.file_portafoglio, sheet_name='Portfolio', header=1)

        ws = self.wb.create_sheet('24.liq_tot')
        ws = self.wb['24.liq_tot']
        self.wb.active = ws

        # Creazione tabella
        header_24 = ['', 'LIQUIDITÀ', 'LIQUIDITÀ IN VALUTA', 'Totale '+ self.mesi_dict[self.t1.month], 'Totale '+ self.mesi_dict[self.t0_1m.month]]
        len_header_24 = len(header_24)

        # Titolo
        ws['A1'] = 'Riepilogo Liquidità'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells(start_row=1, end_row=4, start_column=1, end_column=12)
        min_col = 4

        # Intestazione
        for col in ws.iter_cols(min_row=8, max_row=9, min_col=min_col, max_col=min_col + len_header_24 -1):
            ws[col[0].coordinate].value = header_24[0]
            del header_24[0]
            ws[col[0].coordinate].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col[0].coordinate].font = Font(name='Times New Roman', size=10, color='FFFFFF')
            ws[col[0].coordinate].fill = PatternFill(fill_type='solid', fgColor='31869B')
            ws[col[0].coordinate].border = Border(right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'))
            ws.merge_cells(start_row=col[0].row, end_row=col[1].row, start_column=col[0].column, end_column=col[0].column)
            ws.row_dimensions[col[0].row].height = 20
            ws.row_dimensions[col[1].row].height = 20
            ws.column_dimensions[col[0].column_letter].width = 15

        # Indice e riempimento tabella
        int_liq = list(portfolio.loc[(portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR'),'INTERMEDIARIO'].unique())
        len_int_liq = len(int_liq)
        num_intermediari = 2
        lunghezza_colonna_24 = []
        for row in ws.iter_rows(min_row=8, max_row=10 + len_int_liq -1, min_col=min_col, max_col=min_col + len_header_24):
            if row[0].row > 9:
                ws[row[0].coordinate].value = int_liq[0]
                del int_liq[0]
                ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
                ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, color='000000')
                ws[row[0].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))

                ws[row[1].coordinate].value = portfolio.loc[(portfolio['CATEGORIA']=='CASH') & (portfolio['INTERMEDIARIO']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() if portfolio.loc[(portfolio['CATEGORIA']=='CASH') & (portfolio['INTERMEDIARIO']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() != 0 else ''
                ws[row[1].coordinate].alignment = Alignment(horizontal='center')
                ws[row[1].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[1].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[1].coordinate].number_format = '#,0'
                ws[row[2].coordinate].value = portfolio.loc[(portfolio['CATEGORIA']=='CASH_FOREIGN_CURR') & (portfolio['INTERMEDIARIO']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() if portfolio.loc[(portfolio['CATEGORIA']=='CASH_FOREIGN_CURR') & (portfolio['INTERMEDIARIO']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() != 0 else ''
                ws[row[2].coordinate].alignment = Alignment(horizontal='center')
                ws[row[2].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[2].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[2].coordinate].number_format = '#,0'

                ws[row[num_intermediari+1].coordinate].value = portfolio.loc[(portfolio['INTERMEDIARIO']==ws[row[0].coordinate].value) & ((portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR')), 'TOTALE t1'].sum()
                ws[row[num_intermediari+1].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+1].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+1].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+1].coordinate].number_format = '#,0'
                ws[row[num_intermediari+2].coordinate].value = portfolio.loc[(portfolio['INTERMEDIARIO']==ws[row[0].coordinate].value) & ((portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR')), 'TOTALE t0'].sum()
                ws[row[num_intermediari+2].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+2].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+2].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+2].coordinate].number_format = '#,0'

                lunghezza_colonna_24.append(len(ws.cell(row=row[0].row, column=row[0].column).value)) # ottieni la lunghezza della colonna
                ws.column_dimensions[row[0].column_letter].width = max(lunghezza_colonna_24) + 2.5 # modifica larghezza colonna FALLO ALLA FINE

        # Somma per strumento
        for row in ws.iter_rows(min_row=10 + len_int_liq, max_row=10 + len_int_liq, min_col=min_col, max_col=min_col + len_header_24):
            ws[row[0].coordinate].value = 'TOTALE'
            ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
            ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
            ws[row[0].coordinate].border = Border(bottom=Side(border_style='thin', color='31869B'), right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'), top=Side(border_style='thin', color='31869B'))
            
            ws[row[1].coordinate].value = portfolio.loc[portfolio['CATEGORIA']=='CASH', 'TOTALE t1'].sum()
            ws[row[2].coordinate].value = portfolio.loc[portfolio['CATEGORIA']=='CASH_FOREIGN_CURR', 'TOTALE t1'].sum()
            
            ws[row[len_header_24-2].coordinate].value = portfolio.loc[(portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR'), 'TOTALE t1'].sum()
            ws[row[len_header_24-1].coordinate].value = portfolio.loc[(portfolio['CATEGORIA']=='CASH') | (portfolio['CATEGORIA']=='CASH_FOREIGN_CURR'), 'TOTALE t0'].sum()

            for _ in range(1,len_header_24):
                ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                ws[row[_].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
                ws[row[_].coordinate].border = Border(bottom=Side(border_style='thin', color='31869B'), right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'), top=Side(border_style='thin', color='31869B'))
                ws[row[_].coordinate].number_format = '#,0'

    def gestioni_25(self):
        """
        Crea la venticinquesima pagina.
        """
        # Carica portafoglio
        portfolio = pd.read_excel(self.file_portafoglio, sheet_name='Portfolio', header=1)

        # 25.Gestione
        ws = self.wb.create_sheet('25.ges')
        ws = self.wb['25.ges']
        self.wb.active = ws

        # Creazione tabella
        header_25 = list(portfolio.loc[portfolio['CATEGORIA']=='GP', 'INTERMEDIARIO'].unique())
        header_25.insert(0, '')
        header_25.extend(('Totale '+ self.mesi_dict[self.t1.month], 'Totale '+ self.mesi_dict[self.t0_1m.month], 'Delta'))
        len_header_25 = len(header_25)

        # Titolo
        ws['A1'] = 'Gestioni'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        if len(list(portfolio.loc[portfolio['CATEGORIA']=='GP','INTERMEDIARIO'].unique())) == 1:
            lunghezza_titolo_25 = 12
            min_col = 4
        elif len(list(portfolio.loc[portfolio['CATEGORIA']=='GP','INTERMEDIARIO'].unique())) == 2:
            lunghezza_titolo_25 = 12
            min_col = 4
        elif len(list(portfolio.loc[portfolio['CATEGORIA']=='GP','INTERMEDIARIO'].unique())) == 3:
            lunghezza_titolo_25 = 12
            min_col = 3
        elif len(list(portfolio.loc[portfolio['CATEGORIA']=='GP','INTERMEDIARIO'].unique())) == 4:
            lunghezza_titolo_25 = 12
            min_col = 3
        elif len(list(portfolio.loc[portfolio['CATEGORIA']=='GP','INTERMEDIARIO'].unique())) == 5:
            lunghezza_titolo_25 = 12
            min_col = 2
        elif len(list(portfolio.loc[portfolio['CATEGORIA']=='GP','INTERMEDIARIO'].unique())) == 6:
            lunghezza_titolo_25 = 12
            min_col = 2
        elif len(list(portfolio.loc[portfolio['CATEGORIA']=='GP','INTERMEDIARIO'].unique())) == 7:
            lunghezza_titolo_25 = 12
            min_col = 1
        else:
            lunghezza_titolo_25 = len_header_25
            min_col = 1
        ws.merge_cells(start_row=1, end_row=4, start_column=1, end_column=lunghezza_titolo_25)

        # Intestazione
        for col in ws.iter_cols(min_row=8, max_row=9, min_col=min_col, max_col=min_col + len_header_25 - 1):
            ws[col[0].coordinate].value = header_25[0]
            del header_25[0]
            ws[col[0].coordinate].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col[0].coordinate].font = Font(name='Times New Roman', size=10, color='FFFFFF')
            ws[col[0].coordinate].fill = PatternFill(fill_type='solid', fgColor='31869B')
            ws[col[0].coordinate].border = Border(right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'))
            ws.merge_cells(start_row=col[0].row, end_row=col[1].row, start_column=col[0].column, end_column=col[0].column)
            ws.row_dimensions[col[0].row].height = 20
            ws.row_dimensions[col[1].row].height = 20
            ws.column_dimensions[col[0].column_letter].width = 12

        # Indice e riempimento tabella
        nome_ges = list(portfolio.loc[portfolio['CATEGORIA']=='GP','PRODOTTO'])
        len_nome_ges = len(nome_ges)
        num_intermediari = len(portfolio.loc[portfolio['CATEGORIA']=='GP', 'INTERMEDIARIO'].unique())
        lunghezza_colonna_25 = []
        for row in ws.iter_rows(min_row=8, max_row=10 + len_nome_ges -1, min_col=min_col, max_col=min_col + len_header_25):
            if row[0].row > 9:
                ws[row[0].coordinate].value = nome_ges[0] 
                del nome_ges[0]
                ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
                ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, color='000000')
                ws[row[0].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                for _ in range(1, num_intermediari+1):
                    ws[row[_].coordinate].value = portfolio.loc[(portfolio['INTERMEDIARIO']==ws[row[_].offset(column=0, row=-2-(row[0].row-10)).coordinate].value) & (portfolio['PRODOTTO']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() if portfolio.loc[(portfolio['INTERMEDIARIO']==ws[row[_].offset(column=0, row=-2-(row[0].row-10)).coordinate].value) & (portfolio['PRODOTTO']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() != 0 else ''
                    ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                    ws[row[_].coordinate].font = Font(name='Times New Roman', size=9)
                    ws[row[_].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                    ws[row[_].coordinate].number_format = '#,0'

                ws[row[num_intermediari+1].coordinate].value = portfolio.loc[portfolio['PRODOTTO']==ws[row[0].coordinate].value, 'TOTALE t1'].sum()
                ws[row[num_intermediari+1].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+1].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+1].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+1].coordinate].number_format = '#,0'
                ws[row[num_intermediari+2].coordinate].value = portfolio.loc[portfolio['PRODOTTO']==ws[row[0].coordinate].value, 'TOTALE t0'].sum()
                ws[row[num_intermediari+2].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+2].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+2].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+2].coordinate].number_format = '#,0'
                ws[row[num_intermediari+3].coordinate].value = (ws[row[num_intermediari+1].coordinate].value -  ws[row[num_intermediari+2].coordinate].value) / (ws[row[num_intermediari+2].coordinate].value) if ws[row[num_intermediari+2].coordinate].value != 0 and ws[row[num_intermediari+1].coordinate].value != 0 else '/'
                ws[row[num_intermediari+3].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+3].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+3].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+3].coordinate].number_format = FORMAT_PERCENTAGE_00

                lunghezza_colonna_25.append(len(ws.cell(row=row[0].row, column=row[0].column).value)) # ottieni la lunghezza della colonna
                ws.column_dimensions[row[0].column_letter].width = max(lunghezza_colonna_25) + 3.5 # modifica larghezza colonna 
            
        # Somma per intermediari
        for row in ws.iter_rows(min_row=10 + len_nome_ges, max_row=10 + len_nome_ges, min_col=min_col, max_col=min_col + len_header_25):
            ws[row[0].coordinate].value = 'TOTALE'
            ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
            ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
            ws[row[0].coordinate].border = Border(bottom=Side(border_style='thin', color='31869B'), right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'), top=Side(border_style='thin', color='31869B'))
            for _ in range(1,len_header_25-2):
                ws[row[_].coordinate].value = portfolio.loc[(portfolio['INTERMEDIARIO']==ws.cell(row=row[_].row, column=row[_].column).offset(row=-len_nome_ges-2).value) & ((portfolio['CATEGORIA']=='GP')), 'TOTALE t1'].sum()
            ws[row[len_header_25-3].coordinate].value = portfolio.loc[portfolio['CATEGORIA']=='GP', 'TOTALE t1'].sum()
            ws[row[len_header_25-2].coordinate].value = portfolio.loc[portfolio['CATEGORIA']=='GP', 'TOTALE t0'].sum()
            ws[row[len_header_25-1].coordinate].value = (portfolio.loc[(portfolio['CATEGORIA']=='GP') & (portfolio['TOTALE t0']!=0), 'TOTALE t1'].sum() - portfolio.loc[(portfolio['CATEGORIA']=='GP') & (portfolio['TOTALE t1']!=0), 'TOTALE t0'].sum()) / portfolio.loc[(portfolio['CATEGORIA']=='GP') & (portfolio['TOTALE t1']!=0), 'TOTALE t0'].sum()
            for _ in range(1,len_header_25):
                ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                ws[row[_].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
                ws[row[_].coordinate].border = Border(bottom=Side(border_style='thin', color='31869B'), right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'), top=Side(border_style='thin', color='31869B'))
                ws[row[_].coordinate].number_format = '#,0'
            ws[row[len_header_25-1].coordinate].number_format = FORMAT_PERCENTAGE_00

    def inv_alt_26(self):
        """
        Crea la ventiseiesima pagina.
        """
        # Carica portafoglio
        portfolio = pd.read_excel(self.file_portafoglio, sheet_name='Portfolio', header=1)


        # 26.Inv.Alt
        ws = self.wb.create_sheet('26.invalt')
        ws = self.wb['26.invalt']
        self.wb.active = ws

        # Creazione tabella
        header_26 = list(portfolio.loc[(portfolio['CATEGORIA']=='HEDGE_FUND') | (portfolio['CATEGORIA']=='ALTERNATIVE_ASSET'), 'INTERMEDIARIO'].unique())
        header_26.insert(0, '')
        header_26.extend(('Totale '+ self.mesi_dict[self.t1.month], 'Totale '+ self.mesi_dict[self.t0_1m.month], 'Delta'))
        len_header_26 = len(header_26)

        # Titolo
        ws['A1'] = 'Inv. Alt. e Hedge Fund'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        if len(list(portfolio.loc[(portfolio['CATEGORIA']=='HEDGE_FUND') | (portfolio['CATEGORIA']=='ALTERNATIVE_ASSET'),'INTERMEDIARIO'].unique())) == 1:
            lunghezza_titolo_26 = 12
            min_col = 4
        elif len(list(portfolio.loc[(portfolio['CATEGORIA']=='HEDGE_FUND') | (portfolio['CATEGORIA']=='ALTERNATIVE_ASSET'),'INTERMEDIARIO'].unique())) == 2:
            lunghezza_titolo_26 = 12
            min_col = 4
        elif len(list(portfolio.loc[(portfolio['CATEGORIA']=='HEDGE_FUND') | (portfolio['CATEGORIA']=='ALTERNATIVE_ASSET'),'INTERMEDIARIO'].unique())) == 3:
            lunghezza_titolo_26 = 12
            min_col = 3
        elif len(list(portfolio.loc[(portfolio['CATEGORIA']=='HEDGE_FUND') | (portfolio['CATEGORIA']=='ALTERNATIVE_ASSET'),'INTERMEDIARIO'].unique())) == 4:
            lunghezza_titolo_26 = 12
            min_col = 3
        elif len(list(portfolio.loc[(portfolio['CATEGORIA']=='HEDGE_FUND') | (portfolio['CATEGORIA']=='ALTERNATIVE_ASSET'),'INTERMEDIARIO'].unique())) == 5:
            lunghezza_titolo_26 = 12
            min_col = 2
        elif len(list(portfolio.loc[(portfolio['CATEGORIA']=='HEDGE_FUND') | (portfolio['CATEGORIA']=='ALTERNATIVE_ASSET'),'INTERMEDIARIO'].unique())) == 6:
            lunghezza_titolo_26 = 12
            min_col = 2
        elif len(list(portfolio.loc[(portfolio['CATEGORIA']=='HEDGE_FUND') | (portfolio['CATEGORIA']=='ALTERNATIVE_ASSET'),'INTERMEDIARIO'].unique())) == 7:
            lunghezza_titolo_26 = 12
            min_col = 1
        else:
            lunghezza_titolo_26 = len_header_26
            min_col = 1
        ws.merge_cells(start_row=1, end_row=4, start_column=1, end_column=lunghezza_titolo_26)

        # Intestazione
        for col in ws.iter_cols(min_row=8, max_row=9, min_col=min_col, max_col=min_col + len_header_26 - 1):
            ws[col[0].coordinate].value = header_26[0]
            del header_26[0]
            ws[col[0].coordinate].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col[0].coordinate].font = Font(name='Times New Roman', size=10, color='FFFFFF')
            ws[col[0].coordinate].fill = PatternFill(fill_type='solid', fgColor='31869B')
            ws[col[0].coordinate].border = Border(right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'))
            ws.merge_cells(start_row=col[0].row, end_row=col[1].row, start_column=col[0].column, end_column=col[0].column)
            ws.row_dimensions[col[0].row].height = 20
            ws.row_dimensions[col[1].row].height = 20
            ws.column_dimensions[col[0].column_letter].width = 12

        # Indice e riempimento tabella
        nome_invalt = list(portfolio.loc[(portfolio['CATEGORIA']=='HEDGE_FUND') | (portfolio['CATEGORIA']=='ALTERNATIVE_ASSET'),'PRODOTTO'])
        len_nome_invalt = len(nome_invalt)
        num_intermediari = len(portfolio.loc[(portfolio['CATEGORIA']=='HEDGE_FUND') | (portfolio['CATEGORIA']=='ALTERNATIVE_ASSET'), 'INTERMEDIARIO'].unique())
        lunghezza_colonna_26 = []
        for row in ws.iter_rows(min_row=8, max_row=10 + len_nome_invalt -1, min_col=min_col, max_col=min_col + len_header_26):
            if row[0].row > 9:
                ws[row[0].coordinate].value = nome_invalt[0] 
                del nome_invalt[0]
                ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
                ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, color='000000')
                ws[row[0].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                for _ in range(1, num_intermediari+1):
                    ws[row[_].coordinate].value = portfolio.loc[(portfolio['INTERMEDIARIO']==ws[row[_].offset(column=0, row=-2-(row[0].row-10)).coordinate].value) & (portfolio['PRODOTTO']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() if portfolio.loc[(portfolio['INTERMEDIARIO']==ws[row[_].offset(column=0, row=-2-(row[0].row-10)).coordinate].value) & (portfolio['PRODOTTO']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() != 0 else ''
                    ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                    ws[row[_].coordinate].font = Font(name='Times New Roman', size=9)
                    ws[row[_].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                    ws[row[_].coordinate].number_format = '#,0'

                ws[row[num_intermediari+1].coordinate].value = portfolio.loc[portfolio['PRODOTTO']==ws[row[0].coordinate].value, 'TOTALE t1'].sum()
                ws[row[num_intermediari+1].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+1].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+1].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+1].coordinate].number_format = '#,0'
                ws[row[num_intermediari+2].coordinate].value = portfolio.loc[portfolio['PRODOTTO']==ws[row[0].coordinate].value, 'TOTALE t0'].sum()
                ws[row[num_intermediari+2].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+2].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+2].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+2].coordinate].number_format = '#,0'
                ws[row[num_intermediari+3].coordinate].value = (ws[row[num_intermediari+1].coordinate].value -  ws[row[num_intermediari+2].coordinate].value) / (ws[row[num_intermediari+2].coordinate].value) if ws[row[num_intermediari+2].coordinate].value != 0 and ws[row[num_intermediari+1].coordinate].value != 0 else '/'
                ws[row[num_intermediari+3].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+3].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+3].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+3].coordinate].number_format = FORMAT_PERCENTAGE_00

                lunghezza_colonna_26.append(len(ws.cell(row=row[0].row, column=row[0].column).value)) # ottieni la lunghezza della colonna
                ws.column_dimensions[row[0].column_letter].width = max(lunghezza_colonna_26) + 2.5 # modifica larghezza colonna

        # Somma per intermediari
        for row in ws.iter_rows(min_row=10 + len_nome_invalt, max_row=10 + len_nome_invalt, min_col=min_col, max_col=min_col + len_header_26):
            ws[row[0].coordinate].value = 'TOTALE'
            ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
            ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
            ws[row[0].coordinate].border = Border(bottom=Side(border_style='thin', color='31869B'), right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'), top=Side(border_style='thin', color='31869B'))
            for _ in range(1,len_header_26-2):
                ws[row[_].coordinate].value = portfolio.loc[(portfolio['INTERMEDIARIO']==ws.cell(row=row[_].row, column=row[_].column).offset(row=-len_nome_invalt-2).value) & ((portfolio['CATEGORIA']=='HEDGE_FUND') | (portfolio['CATEGORIA']=='ALTERNATIVE_ASSET')), 'TOTALE t1'].sum()
            ws[row[len_header_26-3].coordinate].value = portfolio.loc[(portfolio['CATEGORIA']=='HEDGE_FUND') | (portfolio['CATEGORIA']=='ALTERNATIVE_ASSET'), 'TOTALE t1'].sum()
            ws[row[len_header_26-2].coordinate].value = portfolio.loc[(portfolio['CATEGORIA']=='HEDGE_FUND') | (portfolio['CATEGORIA']=='ALTERNATIVE_ASSET'), 'TOTALE t0'].sum()
            ws[row[len_header_26-1].coordinate].value = (portfolio.loc[(portfolio['CATEGORIA']=='HEDGE_FUND') | (portfolio['CATEGORIA']=='ALTERNATIVE_ASSET') & (portfolio['TOTALE t0']!=0), 'TOTALE t1'].sum() - portfolio.loc[(portfolio['CATEGORIA']=='HEDGE_FUND') | (portfolio['CATEGORIA']=='ALTERNATIVE_ASSET') & (portfolio['TOTALE t1']!=0), 'TOTALE t0'].sum()) / portfolio.loc[(portfolio['CATEGORIA']=='HEDGE_FUND') | (portfolio['CATEGORIA']=='ALTERNATIVE_ASSET') & (portfolio['TOTALE t1']!=0), 'TOTALE t0'].sum()
            for _ in range(1,len_header_26):
                ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                ws[row[_].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
                ws[row[_].coordinate].border = Border(bottom=Side(border_style='thin', color='31869B'), right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'), top=Side(border_style='thin', color='31869B'))
                ws[row[_].coordinate].number_format = '#,0'
            ws[row[len_header_26-1].coordinate].number_format = FORMAT_PERCENTAGE_00

    def asset_allocation_27(self):
        """
        Crea la ventisettesima pagina.
        """
        # Carica portafoglio
        portfolio = pd.read_excel(self.file_portafoglio, sheet_name='Portfolio', header=1)
        # Carica asset-allocation gestioni
        ass_allocation = pd.read_excel(self.file_portafoglio, sheet_name='Gestioni', header=0)

        # 27.Sintesi
        ws = self.wb.create_sheet('27.ass_all')
        ws = self.wb['27.ass_all']
        self.wb.active = ws

        # Creazione tabella
        header_27 = list(portfolio['INTERMEDIARIO'].unique())
        header_27.insert(0, '')
        header_27.extend(('Totale '+ self.mesi_dict[self.t1.month], 'Totale '+ self.mesi_dict[self.t0_1m.month]))
        len_header_27 = len(header_27)
        # Cerca le posizioni contenenti almeno una gestione patrimoniale
        posizioni_con_gestioni = []
        for intermediario in header_27:
            if 'GP' in portfolio.loc[portfolio['INTERMEDIARIO']==intermediario, 'CATEGORIA'].unique():
                posizioni_con_gestioni.append(intermediario)
        #print(posizioni_con_gestioni)

        # Titolo
        ws['A1'] = 'Asset Allocation'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        if len(list(portfolio['INTERMEDIARIO'].unique())) == 1:
            lunghezza_titolo_27 = 12
            min_col = 4
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 2:
            lunghezza_titolo_27 = 12
            min_col = 4
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 3:
            lunghezza_titolo_27 = 12
            min_col = 3
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 4:
            lunghezza_titolo_27 = 12
            min_col = 3
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 5:
            lunghezza_titolo_27 = 12
            min_col = 2
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 6:
            lunghezza_titolo_27 = 12
            min_col = 2
        elif len(list(portfolio['INTERMEDIARIO'].unique())) == 7:
            lunghezza_titolo_27 = 12
            min_col = 1
        else:
            lunghezza_titolo_27 = len_header_27
            min_col = 1
        ws.merge_cells(start_row=1, end_row=4, start_column=1, end_column=lunghezza_titolo_27)

        for col in ws.iter_cols(min_row=8, max_row=9, min_col=min_col, max_col=min_col + len_header_27 - 1):
            ws[col[0].coordinate].value = header_27[0]
            del header_27[0]
            ws[col[0].coordinate].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col[0].coordinate].font = Font(name='Times New Roman', size=10, color='FFFFFF')
            ws[col[0].coordinate].fill = PatternFill(fill_type='solid', fgColor='31869B')
            ws[col[0].coordinate].border = Border(right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'))
            ws.merge_cells(start_row=col[0].row, end_row=col[1].row, start_column=col[0].column, end_column=col[0].column)
            ws.row_dimensions[col[0].row].height = 20
            ws.row_dimensions[col[1].row].height = 20
            ws.column_dimensions[col[0].column_letter].width = 12

        # tipo_strumento_nogp = list(portfolio.loc[portfolio['CATEGORIA']!='GP', 'CATEGORIA'].unique())
        # Con la rimozione dell'hedge fund da JPM, non esistono più prodotti di quel tipo, se non all'interno
        # delle gestioni patrimoniali. Con la riga di codice sopra, non vengono intercettati
        # Chiedi al professor Maspero di poter togliere dall'asset allocation gli hedge fund.
        # Devo specificarli tutti a mano
        tipo_strumento_nogp = [
            'CASH', 'EQUITY', 'CASH_FOREIGN_CURR', 'CORPORATE_BOND', 'GOVERNMENT_BOND', 
            'ALTERNATIVE_ASSET', 'HEDGE_FUND'
        ]
        len_tipo_strumento_nogp = len(tipo_strumento_nogp)
        num_intermediari = len(portfolio['INTERMEDIARIO'].unique())
        lunghezza_colonna_27 = []
        tipo_strumento_dict = {
            'CASH' : 'LIQUIDITÀ', 'GP' : 'GESTIONI', 'EQUITY' : 'AZIONI', 'CASH_FOREIGN_CURR' : 'LIQUIDITÀ IN VALUTA', 
            'CORPORATE_BOND' : 'OBBLIGAZIONI CORPORATE', 'GOVERNMENT_BOND' : 'OBBLIGAZIONI GOVERNATIVE', 
            'ALTERNATIVE_ASSET' : 'INVESTIMENTI ALTERNATIVI', 'HEDGE_FUND' : 'HEDGE FUND'
        }
        for row in ws.iter_rows(min_row=8, max_row=10 + len_tipo_strumento_nogp -1, min_col=min_col, max_col=min_col + len_header_27):
            if row[0].row > 9:
                ws[row[0].coordinate].value = tipo_strumento_nogp[0] # carica i tipi di strumenti nell'indice
                del tipo_strumento_nogp[0]
                ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
                ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, color='000000')
                ws[row[0].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws.row_dimensions[row[0].row].height = 19

                for _ in range(1, num_intermediari+1):
                    ws[row[_].coordinate].value = portfolio.loc[(portfolio['INTERMEDIARIO']==ws[row[_].offset(column=0, row=-2-(row[0].row-10)).coordinate].value) & (portfolio['CATEGORIA']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() if portfolio.loc[(portfolio['INTERMEDIARIO']==ws[row[_].offset(column=0, row=-2-(row[0].row-10)).coordinate].value) & (portfolio['CATEGORIA']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() != 0 else ''
                    ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                    ws[row[_].coordinate].font = Font(name='Times New Roman', size=9)
                    ws[row[_].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                    ws[row[_].coordinate].number_format = '#,0'

                for _ in range(1, num_intermediari+1):
                    if ws[row[_].offset(column=0, row=-2-(row[0].row-10)).coordinate].value in posizioni_con_gestioni:
                        ws[row[_].coordinate].value = ass_allocation.loc[(ass_allocation['INTERMEDIARIO']==ws[row[_].offset(column=0, row=-2-(row[0].row-10)).coordinate].value) & (ass_allocation['CATEGORIA']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() if ass_allocation.loc[(ass_allocation['INTERMEDIARIO']==ws[row[_].offset(column=0, row=-2-(row[0].row-10)).coordinate].value) & (ass_allocation['CATEGORIA']==ws[row[0].coordinate].value), 'TOTALE t1'].sum() != 0 else ''
                
                ws[row[num_intermediari+1].coordinate].value = portfolio.loc[(portfolio['CATEGORIA']==ws[row[0].coordinate].value) & (~portfolio['INTERMEDIARIO'].isin(posizioni_con_gestioni)), 'TOTALE t1'].sum() + ass_allocation.loc[(ass_allocation['CATEGORIA']==ws[row[0].coordinate].value) & (ass_allocation['INTERMEDIARIO'].isin(posizioni_con_gestioni)), 'TOTALE t1'].sum()
                ws[row[num_intermediari+1].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+1].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+1].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+1].coordinate].number_format = '#,0'
                # TODO : controlla la somma dei valori nella colonna totale mese t0 usando valori veri.
                ws[row[num_intermediari+2].coordinate].value = portfolio.loc[(portfolio['CATEGORIA']==ws[row[0].coordinate].value) & (~portfolio['INTERMEDIARIO'].isin(posizioni_con_gestioni)), 'TOTALE t0'].sum() + ass_allocation.loc[(ass_allocation['CATEGORIA']==ws[row[0].coordinate].value) & (ass_allocation['INTERMEDIARIO'].isin(posizioni_con_gestioni)), 'TOTALE t0'].sum()
                ws[row[num_intermediari+2].coordinate].alignment = Alignment(horizontal='center')
                ws[row[num_intermediari+2].coordinate].font = Font(name='Times New Roman', size=9)
                ws[row[num_intermediari+2].coordinate].border = Border(bottom=Side(border_style='dashed', color='31869B'), right=Side(border_style='dashed', color='31869B'), left=Side(border_style='dashed', color='31869B'))
                ws[row[num_intermediari+2].coordinate].number_format = '#,0'

                ws[row[0].coordinate].value = tipo_strumento_dict[ws[row[0].coordinate].value] # aggiorna valori dell'indice con i nomi nel dizionario
                lunghezza_colonna_27.append(len(ws.cell(row=row[0].row, column=row[0].column).value)) # ottieni la lunghezza della colonna
                ws.column_dimensions[row[0].column_letter].width = max(lunghezza_colonna_27) + 2.5 # modifica larghezza colonna

        # Somma per intermediari
        for row in ws.iter_rows(min_row=10 + len_tipo_strumento_nogp, max_row=10 + len_tipo_strumento_nogp, min_col=min_col, max_col=min_col + len_header_27):
            ws[row[0].coordinate].value = 'TOTALE'
            ws[row[0].coordinate].alignment = Alignment(horizontal='left', vertical='center')
            ws[row[0].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
            ws[row[0].coordinate].border = Border(bottom=Side(border_style='thin', color='31869B'), right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'), top=Side(border_style='thin', color='31869B'))
            ws.row_dimensions[row[0].row].height = 19
            for _ in range(1,len_header_27-2):
                ws[row[_].coordinate].value = portfolio.loc[portfolio['INTERMEDIARIO']==ws.cell(row=row[_].row, column=row[_].column).offset(row=-len_tipo_strumento_nogp-2).value, 'TOTALE t1'].sum()
            ws[row[len_header_27-2].coordinate].value = portfolio.loc[:, 'TOTALE t1'].sum()
            ws[row[len_header_27-1].coordinate].value = portfolio.loc[:, 'TOTALE t0'].sum()
            for _ in range(1,len_header_27):
                ws[row[_].coordinate].alignment = Alignment(horizontal='center')
                ws[row[_].coordinate].font = Font(name='Times New Roman', size=9, bold=True)
                ws[row[_].coordinate].border = Border(bottom=Side(border_style='thin', color='31869B'), right=Side(border_style='thin', color='31869B'), left=Side(border_style='thin', color='31869B'), top=Side(border_style='thin', color='31869B'))
                ws[row[_].coordinate].number_format = '#,0'

        chart = PieChart()
        labels = Reference(ws, min_col=min_col, max_col=min_col, min_row=10, max_row=10+len_tipo_strumento_nogp-1)
        data = Reference(
            ws, min_col=min_col + len_header_27 - 2, max_col=min_col + len_header_27 - 2, min_row=10, max_row=10+len_tipo_strumento_nogp-1
        )
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(labels)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.textProperties = RichText(
            p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=1200, b=True)),
            endParaRPr=CharacterProperties(sz=1200, b=True))])
        chart.legend.layout = Layout(manualLayout=ManualLayout(h=1))
        palette = ['540b0e', 'ff595e', 'ffca3a', '8ac926', '1982c4', 'f15bb5', '6a4c93']
        series = chart.series[0]
        for point in range(len_tipo_strumento_nogp):
            pt = DataPoint(idx=point)
            pt.graphicalProperties.solidFill = palette[point]
            series.dPt.append(pt)

        ws.add_chart(chart, 'D20')

    def contatti_28(self):
        """
        Crea la ventottesima pagina.
        """
        ws = self.wb.create_sheet('28.contatti')
        ws = self.wb['28.contatti']
        self.wb.active = ws

        ws['A1'] = '4. Contatti'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells('A1:L4')

        header_28 = ['Benchmark & Style S.r.l.', 'Via San Siro, 33', '20149 Milano', '="+390258328666"', 'info@benchmarkandstyle.com']
        for row in ws.iter_rows(min_row=6, max_row=10, min_col=1, max_col=12):
            ws[row[0].coordinate].value = header_28[0]
            del header_28[0]
            ws[row[0].coordinate].alignment = Alignment(horizontal='center', vertical='center')
            ws[row[0].coordinate].font = Font(name='Times New Roman', size=11, bold=True, color='31869B')
            ws.merge_cells(start_row=row[0].row, end_row=row[0].row, start_column=1, end_column=12)

        ws['A13'] = 'Disclaimer'
        ws['A13'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A13'].font = Font(name='Times New Roman', size=48, bold=True, color='FFFFFF')
        ws['A13'].fill = PatternFill(fill_type='solid', fgColor='31869B')
        ws.merge_cells('A13:L16')

        ws['A18'] = 'Il presente rendiconto ha una funzione meramente informativa ed è stato redatto sulla base dei dati forniti dai singoli gestori cui è affidato il patrimonio del cliente. I dati sono stati rielaborati al fine di fornire una visione d\'insieme e progressiva dei rendimenti mensili del patrimonio e delle singole gestioni confrontati ai relativi benchmarks. Tale rielaborazione rende più semplice comprendere i contributi dei singoli gestori e delle varie classi d\'attivo alla performance del patrimonio nel periodo considerato nonchè di monitorare periodicamente la performance stessa e i rischi assunti a livello consolidato e dei singoli portafogli.'
        ws['A18'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws['A18'].font = Font(name='Times New Roman', size=11, bold=True, italic=True, color='31869B')
        ws.merge_cells('A18:L24')

        ws['A25'] = 'Il presente rendiconto non rappresenta in alcun caso una raccomandazione e/o sollecitazione all\'acquisto o alla vendita di titoli, fondi, strumenti finanziari derivati, valute o altro e gli eventuali contenuti in esso non potranno in nessun caso essere ritenuti responsabili delle future performance del patrimonio del cliente neppure con riferimento alle previsioni formulate circa la prevedibile evoluzione dei mercati finanziari e/o di singoli comparti di essi.'
        ws['A25'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws['A25'].font = Font(name='Times New Roman', size=11, bold=True, italic=True, color='31869B')
        ws.merge_cells('A25:L28')

        self.__logo(ws)

    def layout(self):
        """
        Cambia il layout dei fogli non nascosti: A4, orizzontale, centrato orizzontalmente, con margini personalizzati, 
        adatta alla pagina oriz. e vert.
        Aggiunge il numero di pagina.
        """
        # Modifica layout di pagina e aggiungi numero di pagina a tutti i fogli
        for sheet in self.wb:
            if sheet.sheet_state != 'hidden':
                self.wb.active = sheet
                sheet.set_printer_settings(paper_size=9, orientation='landscape') # 9 = 'A4'
                sheet.page_margins = PageMargins(left=0.2362204724, right=0.2362204724, top=0.7480314961, bottom=0.0480314961, header=0.3149606299, footer=0.3149606299) #margini personalizzati
                sheet.print_options.horizontalCentered = True # centrato orizzontalmente
                sheet.sheet_properties.pageSetUpPr.fitToPage = True 
                sheet.page_setup.fitToHeight = 1
                sheet.page_setup.fitToWidth = 1
                numero_pagina_regex = re.compile(r'\d*') # regex per trovare il numero del foglio
                numero_pagina_search = numero_pagina_regex.search(str(sheet.title))
                numero_pagina = numero_pagina_search.group()
                sheet.oddFooter.right.text = numero_pagina # assegna il numero del foglio al piè di pagina destro

    def salva_file(self):
        """Salva il file excel.
        """
        self.wb.save(self.path.joinpath('report.xlsx'))


if __name__ == "__main__":
    start = time.time() # TODO: sostituisci tutte le chiamate al foglio Portfolio con Portfolio (2)
    _ = Report(t1='31/12/2024')
    _.copertina_1()
    _.indice_2()
    _.analisi_di_mercato_3()
    _.analisi_rendimenti_4()
    _.analisi_indici_5()
    _.performance_6()
    _.andamento_7()
    _.caricamento_dati()
    _.cono_8()
    _.cono_9()
    _.nuovo_bk_10()
    _.performance_11()
    _.prezzi_12()
    _.prezzi_13()
    _.prezzi_14()
    _.att_in_corso_15()
    _.valutazione_per_macroclasse_16()
    _.sintesi_17()
    _.valuta_18()
    _.tabella_pivot_azioni()
    _.tabella_pivot_obbligazioni_governative()
    _.tabella_pivot_obbligazioni_societarie()
    _.obb_totale_22()
    _.liquidità_23()
    _.liq_totale_24()
    _.gestioni_25()
    _.inv_alt_26()
    _.asset_allocation_27()
    _.contatti_28()
    _.layout()
    _.salva_file()
    end = time.time()
    print("Elapsed time : ", round(end - start, 2), 'seconds')